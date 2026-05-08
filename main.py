from fastapi import FastAPI, UploadFile, File, Form, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from pathlib import Path
import pandas as pd
import json
import re
import unicodedata
import logging
import hashlib
import time
import traceback
import os
import math
from rapidfuzz import fuzz
from html import escape
from typing import Optional, Dict, Any, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

RUBRICS_DIR = BASE_DIR / "rubrics"
RUBRICS_DIR.mkdir(exist_ok=True)

LEGACY_RUBRIC_PATH = BASE_DIR / "rubric_psicolinguistica_2026.json"

# ============================================================
# ROBUSTEZ TÉCNICA + EMBEDDINGS v3.5: BASELINE VALIDACIÓN, EMBEDDINGS OPTIMIZADOS, FALLBACK, CACHÉ Y TRAZABILIDAD
# ============================================================

APP_VERSION = "4.1.0-ocr-v1.1"
LOG_PATH = OUTPUT_DIR / "evalia_runtime.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("evalia")

app = FastAPI(title="Evalia OCR-MVP", version="4.1.0-ocr-v1.1")

SEMANTIC_CACHE: Dict[str, Any] = {}

# ============================================================
# EMBEDDINGS SEMÁNTICOS v3.5 — activación real segura
# ============================================================
# Evalia puede usar embeddings reales si el entorno tiene instalado
# sentence-transformers y el modelo está disponible. Si no, la app
# sigue funcionando con el motor semántico por reglas de v3.1/v3.2.
#
# Para activar embeddings reales en Render/local:
#   1) agregar sentence-transformers a requirements.txt
#   2) definir EVALIA_EMBEDDINGS=1
#   3) opcional: EVALIA_EMBEDDING_MODEL=paraphrase-multilingual-MiniLM-L12-v2
#   4) opcional: EVALIA_EMBEDDING_DEVICE=cpu

EMBEDDINGS_ENABLED = os.getenv("EVALIA_EMBEDDINGS", "0").strip().lower() in {"1", "true", "yes", "on"}
EMBEDDING_MODEL_NAME = os.getenv("EVALIA_EMBEDDING_MODEL", "paraphrase-multilingual-MiniLM-L12-v2")
EMBEDDING_DEVICE = os.getenv("EVALIA_EMBEDDING_DEVICE", "cpu")
EMBEDDING_MODEL = None
EMBEDDING_STATUS = {
    "enabled_requested": EMBEDDINGS_ENABLED,
    "available": False,
    "model": EMBEDDING_MODEL_NAME,
    "device": EMBEDDING_DEVICE,
    "mode": "reglas_semanticas",
    "message": "Embeddings no activados; usando motor semántico por reglas."
}

def get_embedding_model():
    global EMBEDDING_MODEL
    if not EMBEDDINGS_ENABLED:
        return None
    if EMBEDDING_MODEL is not None:
        return EMBEDDING_MODEL
    try:
        from sentence_transformers import SentenceTransformer
        EMBEDDING_MODEL = SentenceTransformer(EMBEDDING_MODEL_NAME, device=EMBEDDING_DEVICE)
        EMBEDDING_STATUS.update({
            "available": True,
            "mode": "hibrido_reglas_embeddings",
            "message": f"Embeddings reales activos con modelo {EMBEDDING_MODEL_NAME} en dispositivo {EMBEDDING_DEVICE}."
        })
        log_event("embedding_model_loaded", model=EMBEDDING_MODEL_NAME, device=EMBEDDING_DEVICE)
        return EMBEDDING_MODEL
    except Exception as e:
        EMBEDDING_STATUS.update({
            "available": False,
            "mode": "reglas_semanticas",
            "message": f"Embeddings solicitados, pero no disponibles: {str(e)}"
        })
        log_event("embedding_model_unavailable", error=str(e))
        return None

def cosine_similarity(vec_a, vec_b):
    try:
        import numpy as np
        a = np.asarray(vec_a, dtype=float)
        b = np.asarray(vec_b, dtype=float)
        denom = (np.linalg.norm(a) * np.linalg.norm(b))
        if denom == 0:
            return 0.0
        return float(np.dot(a, b) / denom)
    except Exception:
        return 0.0

def embedding_vector(text):
    txt = normalize_text(text)
    if not txt:
        return None
    key = cache_key("embedding_vector", txt, EMBEDDING_MODEL_NAME)
    if key in SEMANTIC_CACHE:
        return SEMANTIC_CACHE[key]
    model = get_embedding_model()
    if model is None:
        return None
    try:
        vec = model.encode(txt, normalize_embeddings=True)
        if len(SEMANTIC_CACHE) < 30000:
            SEMANTIC_CACHE[key] = vec
        return vec
    except Exception as e:
        EMBEDDING_STATUS.update({
            "available": False,
            "mode": "reglas_semanticas",
            "message": f"Fallo durante cálculo de embedding: {str(e)}"
        })
        return None

def precompute_embedding_vectors(texts):
    """
    Precalcula embeddings en lote para acelerar el procesamiento.
    Mantiene fallback seguro: si el modelo no está disponible o falla, Evalia sigue con reglas.
    """
    if not EMBEDDINGS_ENABLED:
        return {"requested": False, "precomputed": 0, "seconds": 0.0, "status": "disabled"}

    model = get_embedding_model()
    if model is None:
        return {"requested": True, "precomputed": 0, "seconds": 0.0, "status": "model_unavailable"}

    started = time.time()
    unique_texts = []
    seen = set()
    for t in texts:
        txt = normalize_text(t)
        if not txt or txt in seen:
            continue
        key = cache_key("embedding_vector", txt, EMBEDDING_MODEL_NAME)
        if key in SEMANTIC_CACHE:
            continue
        seen.add(txt)
        unique_texts.append(txt)

    if not unique_texts:
        return {"requested": True, "precomputed": 0, "seconds": 0.0, "status": "cache_hit"}

    try:
        batch_size = int(os.getenv("EVALIA_EMBEDDING_BATCH_SIZE", "32"))
        vectors = model.encode(
            unique_texts,
            normalize_embeddings=True,
            batch_size=batch_size,
            show_progress_bar=False
        )
        stored = 0
        for txt, vec in zip(unique_texts, vectors):
            if len(SEMANTIC_CACHE) >= 50000:
                break
            SEMANTIC_CACHE[cache_key("embedding_vector", txt, EMBEDDING_MODEL_NAME)] = vec
            stored += 1
        seconds = round(time.time() - started, 3)
        EMBEDDING_STATUS.update({
            "available": True,
            "mode": "hibrido_reglas_embeddings_batch",
            "message": f"Embeddings reales activos y optimizados por lote. Modelo {EMBEDDING_MODEL_NAME}."
        })
        log_event("embedding_batch_precomputed", items=stored, seconds=seconds)
        return {"requested": True, "precomputed": stored, "seconds": seconds, "status": "ok"}
    except Exception as e:
        EMBEDDING_STATUS.update({
            "available": False,
            "mode": "reglas_semanticas",
            "message": f"Fallo durante precálculo batch de embeddings: {str(e)}"
        })
        log_event("embedding_batch_failed", error=str(e))
        return {"requested": True, "precomputed": 0, "seconds": round(time.time() - started, 3), "status": "failed"}


def collect_embedding_texts(df, rubric):
    """Recolecta textos frecuentes para precalcular embeddings antes del loop principal."""
    texts = []
    for q in rubric.get("questions", []):
        texts.append(q.get("prompt", ""))
        for item in q.get("accepted_answers", []) or []:
            texts.append(item)
            texts.extend(synonym_expansions(item))
        for c in q.get("criteria", []) or []:
            concept = c.get("concept", "")
            texts.append(concept)
            texts.extend(c.get("semantic_variants", []) or [])
            texts.extend(c.get("accepted_values", []) or [])
            texts.extend(synonym_expansions(concept))
        for pair in q.get("pairs", []) or []:
            texts.append(pair.get("prompt_value", ""))
            texts.append(pair.get("correct_match", ""))
            texts.extend(synonym_expansions(pair.get("prompt_value", "")))
            texts.extend(synonym_expansions(pair.get("correct_match", "")))

        col = find_item_column(df, q.get("id"))
        if col:
            for val in df[col].fillna("").tolist():
                texts.append(val)
    return texts


def embedding_similarity_score(answer, target):
    key = cache_key("embedding_similarity", answer, target, EMBEDDING_MODEL_NAME)
    if key in SEMANTIC_CACHE:
        return SEMANTIC_CACHE[key]
    va = embedding_vector(answer)
    vt = embedding_vector(target)
    if va is None or vt is None:
        return None
    sim = cosine_similarity(va, vt)
    # Convierte similitud coseno a escala 0-100. En respuestas breves, 0.55-0.75 suele ser útil.
    score = max(0, min(100, round(((sim + 1) / 2) * 100)))
    result = (score, round(sim, 4))
    if len(SEMANTIC_CACHE) < 30000:
        SEMANTIC_CACHE[key] = result
    return result

def cache_key(*parts):
    raw = "||".join(str(p) for p in parts)
    return hashlib.md5(raw.encode("utf-8")).hexdigest()

def log_event(event, **kwargs):
    try:
        payload = " | ".join(f"{k}={v}" for k, v in kwargs.items())
        logger.info(f"{event} | {payload}" if payload else event)
    except Exception:
        pass

def safe_error_page(title, message, detail=None, badge="CRB Engine · v3.5 baseline validación"):
    detail_html = f"<br><small>{escape(str(detail))}</small>" if detail else ""
    return HTMLResponse(
        f'''
        <!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Error · Evalia</title>{base_css()}</head>
        <body><div class="page"><main class="shell">{shell_topbar("Error controlado", badge)}<div class="result-card">
          <h1>{escape(title)}</h1>
          <div class="error"><strong>{escape(message)}</strong>{detail_html}</div>
          <p class="lead">Evalia detuvo el procesamiento para evitar un reporte inconsistente. Revisa el formato o vuelve a intentar con el modelo descargable.</p>
          <a class="button" href="/">Volver</a>
        </div>{footer_altiora()}</main></div></body></html>
        ''',
        status_code=400
    )

@app.exception_handler(Exception)
async def evalia_unhandled_exception_handler(request: Request, exc: Exception):
    logger.error("unhandled_exception | path=%s | error=%s\n%s", request.url.path, exc, traceback.format_exc())
    return safe_error_page(
        "Error interno controlado",
        "Ocurrió un problema durante el procesamiento, pero fue registrado para depuración.",
        str(exc)
    )


# ============================================================
# NORMALIZACIÓN GENERAL
# ============================================================

def normalize_text(text):
    if pd.isna(text):
        return ""
    text = str(text).strip().lower()
    text = "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )
    text = re.sub(r"[^a-z0-9ñ\s/._-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def split_values(value):
    """
    Separador flexible para docentes.
    Acepta punto y coma, coma, barra, saltos de línea y viñetas simples.
    """
    if pd.isna(value) or value is None:
        return []
    txt = str(value).strip()
    if not txt:
        return []

    txt = txt.replace("•", "\n").replace("·", "\n")
    txt = re.sub(r"\s+y\s+", ";", txt, flags=re.IGNORECASE)

    # Matching con ':' conserva pares cuando están separados por ; o saltos.
    if ":" in txt:
        parts = re.split(r"\s*(?:;|\n|/|\|)\s*", txt)
    else:
        parts = re.split(r"\s*(?:;|\n|/|\||,)\s*", txt)

    cleaned = []
    for p in parts:
        p = re.sub(r"^\s*[-–—\d\)\.]+\s*", "", p).strip()
        if p:
            cleaned.append(p)
    return cleaned


def normalize_item_type(value):
    t = normalize_text(value).replace(" ", "_")

    mapping = {
        "vf": "true_false",
        "v_f": "true_false",
        "verdadero_falso": "true_false",
        "verdadero/falso": "true_false",
        "verdaderofalso": "true_false",
        "true_false": "true_false",
        "truefalse": "true_false",

        "completacion": "completion",
        "completación": "completion",
        "completar": "completion",
        "complete": "completion",
        "completion": "completion",
        "exacta": "short_exact_answer",
        "respuesta_exacta": "short_exact_answer",
        "short_exact_answer": "short_exact_answer",

        "enumeracion": "enumeration_conceptual",
        "enumeración": "enumeration_conceptual",
        "enumerar": "enumeration_conceptual",
        "lista": "enumeration_conceptual",
        "listado": "enumeration_conceptual",
        "mencionar": "enumeration_conceptual",
        "enumeration": "enumeration_conceptual",
        "enumeration_conceptual": "enumeration_conceptual",
        "lista": "enumeration_conceptual",

        "matching": "classification_matching",
        "emparejamiento": "classification_matching",
        "relacion": "classification_matching",
        "relación": "classification_matching",
        "relacionar": "classification_matching",
        "relaciona": "classification_matching",
        "pares": "classification_matching",
        "une": "classification_matching",
        "classification_matching": "classification_matching",

        "criterios": "criteria",
        "criterio": "criteria",
        "abierta": "criteria",
        "pregunta_abierta": "criteria",
        "preguntaabierta": "criteria",
        "desarrollo": "criteria",
        "explicacion": "criteria",
        "explicación": "criteria",
        "respuesta_abierta": "criteria",
        "desarrollo": "criteria",
        "ensayo": "criteria",
        "explicacion": "criteria",
        "explicación": "criteria",
        "criteria": "criteria",
    }

    return mapping.get(t, t or "criteria")


def normalize_column_name(name):
    n = normalize_text(name)
    n = n.replace(" ", "")
    n = n.replace("-", "_")
    n = n.replace(".", "_")
    return n



def display_item_type(item_type):
    mapping = {
        "criteria": "Pregunta abierta",
        "true_false": "Verdadero/Falso",
        "completion": "Completar",
        "short_exact_answer": "Respuesta breve",
        "enumeration_conceptual": "Enumerar",
        "enumeration_closed": "Enumerar",
        "enumeration_categorized": "Enumerar",
        "enumeration": "Enumerar",
        "classification_matching": "Relacionar"
    }
    return mapping.get(item_type, str(item_type or "Pregunta"))


def normalize_teacher_headers_for_output(name):
    mapping = {
        "pregunta": "Pregunta",
        "tipo_item": "Tipo de pregunta",
        "puntaje_maximo": "Puntaje máximo",
        "promedio_puntaje": "Promedio puntaje",
        "promedio_porcentaje": "Promedio %",
        "confianza_promedio": "Confianza promedio",
        "aceptacion_pct": "Alta confianza %",
        "cautela_pct": "Parcial/intermedia %",
        "revision_pct": "Revisión sugerida %",
        "clasificacion_evalia": "Lectura Evalia",
        "sugerencia_docente": "Sugerencia docente",
        "student_id": "ID estudiante",
        "nombre": "Nombre",
        "porcentaje": "Porcentaje",
        "nivel_desempeno": "Nivel de desempeño",
        "status": "Estado",
        "confidence": "Confianza",
        "feedback": "Retroalimentación",
        "answer": "Respuesta",
        "prompt": "Enunciado",
        "score": "Puntaje",
        "max_score": "Puntaje máximo",
        "pregunta_id": "Pregunta"
    }
    return mapping.get(name, name)


def item_number(item_id):
    m = re.search(r"(\d+)", str(item_id))
    return m.group(1) if m else None


def item_aliases(item_id):
    """
    Permite que una pregunta definida como P1 pueda calzar con Q1,
    pregunta_1, pregunta1, item1, item_1, etc.
    """
    raw = str(item_id).strip()
    n = item_number(raw)

    aliases = {raw, raw.upper(), raw.lower(), normalize_column_name(raw)}

    if n:
        aliases.update({
            f"P{n}", f"p{n}",
            f"Q{n}", f"q{n}",
            f"pregunta{n}", f"pregunta_{n}",
            f"item{n}", f"item_{n}",
            f"i{n}", f"I{n}",
        })

    return {normalize_column_name(a) for a in aliases}


def find_column(df, desired_name, extra_aliases=None):
    normalized_cols = {normalize_column_name(c): c for c in df.columns}
    aliases = {normalize_column_name(desired_name)}

    if extra_aliases:
        aliases.update(normalize_column_name(a) for a in extra_aliases)

    for a in aliases:
        if a in normalized_cols:
            return normalized_cols[a]

    return None


def find_item_column(df, item_id):
    normalized_cols = {normalize_column_name(c): c for c in df.columns}
    for alias in item_aliases(item_id):
        if alias in normalized_cols:
            return normalized_cols[alias]
    return None


def get_student_id(row, df):
    col = find_column(df, "student_id", ["id", "alumno_id", "estudiante_id", "rut", "codigo", "código"])
    return row.get(col, "") if col else ""


def get_student_name(row, df):
    col = find_column(df, "nombre", ["name", "student_name", "alumno", "estudiante", "nombre_estudiante"])
    return row.get(col, "") if col else ""


# ============================================================
# RÚBRICAS JSON Y EXCEL
# ============================================================

def get_available_rubrics():
    rubrics = []

    for path in sorted(RUBRICS_DIR.glob("*.json")):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)

            rubric_name = data.get("name") or data.get("title") or path.stem.replace("_", " ").title()
            rubrics.append({
                "filename": path.name,
                "name": rubric_name,
                "path": path
            })
        except Exception:
            continue

    if not rubrics and LEGACY_RUBRIC_PATH.exists():
        try:
            with open(LEGACY_RUBRIC_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            rubric_name = data.get("name") or data.get("title") or "Psicolingüística 2026"
            rubrics.append({
                "filename": LEGACY_RUBRIC_PATH.name,
                "name": rubric_name,
                "path": LEGACY_RUBRIC_PATH
            })
        except Exception:
            pass

    return rubrics


def safe_rubric_path(selected_filename):
    for r in get_available_rubrics():
        if r["filename"] == selected_filename:
            return r["path"]
    return None


def load_rubric_from_json_path(path):
    with open(path, "r", encoding="utf-8") as f:
        rubric = json.load(f)

    rubric["_rubric_filename"] = path.name
    rubric["_rubric_display_name"] = rubric.get("name") or rubric.get("title") or path.stem.replace("_", " ").title()
    return rubric


def load_selected_rubric(selected_filename=None):
    available = get_available_rubrics()

    if not available:
        raise FileNotFoundError(
            "No se encontraron rúbricas JSON en /rubrics ni rúbrica antigua en la raíz."
        )

    if selected_filename:
        path = safe_rubric_path(selected_filename)
        if path is None:
            raise FileNotFoundError("El Modelo de Rúbrica no existe o no está permitido.")
    else:
        path = available[0]["path"]

    return load_rubric_from_json_path(path)


def build_question_from_excel_row(row):
    """
    Formato simple de rúbrica Excel esperado:

    pregunta | tipo | max_score | respuestas | criterios | required_items | prompt

    Ejemplos:
    P1 | criterios | 3 | lenguaje; mente; psicolingüística | lenguaje; mente; psicolingüística | |
    P2 | VF | 1 | V | | |
    P3 | enumeracion | 2 | memoria; atención; comprensión | | 2 |
    P4 | matching | 3 | Broca:producción; Wernicke:comprensión | | |
    """
    qid = str(row.get("pregunta", "")).strip() or str(row.get("id", "")).strip()
    item_type = normalize_item_type(row.get("tipo", row.get("item_type", "criteria")))
    max_score = float(row.get("max_score", row.get("puntaje", row.get("puntaje_maximo", 1))) or 1)
    prompt = "" if pd.isna(row.get("prompt", "")) else str(row.get("prompt", ""))

    respuesta_simple = row.get("respuesta_esperada", row.get("respuestas", row.get("accepted_answers", "")))
    respuestas = split_values(respuesta_simple)
    criterios = split_values(row.get("criterios", row.get("criteria", "")))
    variantes_semanticas = split_values(row.get("variantes_semanticas", row.get("semantic_variants", "")))

    if item_type == "criteria" and not criterios:
        criterios = respuestas

    question = {
        "id": qid,
        "item_type": item_type,
        "max_score": max_score,
        "prompt": prompt
    }

    if item_type == "true_false":
        question["accepted_answers"] = respuestas or ["V", "Verdadero"]

    elif item_type in ["completion", "short_exact_answer"]:
        question["accepted_answers"] = respuestas

    elif item_type in ["enumeration", "enumeration_conceptual", "enumeration_closed", "enumeration_categorized"]:
        question["accepted_answers"] = respuestas
        required = row.get("required_items", row.get("required_number_of_items", ""))
        if not pd.isna(required) and str(required).strip():
            question["constraints"] = {"required_number_of_items": int(float(required))}

    elif item_type == "classification_matching":
        pairs = []
        for pair_txt in respuestas:
            if ":" in pair_txt:
                left, right = pair_txt.split(":", 1)
            elif "->" in pair_txt:
                left, right = pair_txt.split("->", 1)
            else:
                continue
            pairs.append({
                "prompt_value": left.strip(),
                "correct_match": right.strip(),
                "weight": max_score / max(len(respuestas), 1)
            })
        question["pairs"] = pairs

    else:
        # criteria / abierta
        concepts = criterios or respuestas
        criteria = []
        weight = max_score / max(len(concepts), 1)
        for c in concepts:
            expanded_variants = [c] + variantes_semanticas
            criteria.append({
                "concept": c,
                "weight": weight,
                "semantic_variants": expanded_variants,
                "accepted_values": []
            })
        question["item_type"] = "criteria"
        question["criteria"] = criteria

    return question


def load_rubric_from_excel(path):
    df = pd.read_excel(path)

    # Normalizar encabezados para aceptar nombres simples.
    original_cols = list(df.columns)
    col_map = {}
    for c in original_cols:
        nc = normalize_column_name(c)
        if nc in ["pregunta", "id", "item", "item_id"]:
            col_map[c] = "pregunta"
        elif nc in ["tipo", "tipodepregunta", "tipo_pregunta", "item_type", "tipo_item"]:
            col_map[c] = "tipo"
        elif nc in ["max_score", "puntaje", "puntajemaximo", "puntaje_maximo", "puntajeitem", "puntos"]:
            col_map[c] = "max_score"
        elif nc in ["respuestas", "accepted_answers", "respuesta", "respuestascorrectas", "respuesta_correcta", "respuestaesperada", "respuesta_esperada", "esperado", "esperada", "ideasesperadas", "ideas_esperadas", "ideasclave", "ideas_clave"]:
            col_map[c] = "respuestas"
        elif nc in ["criterios", "criteria", "conceptos", "concepts", "ideasclave", "ideas_clave", "conceptosclave", "conceptos_clave"]:
            col_map[c] = "criterios"
        elif nc in ["variantes", "variantessemanticas", "variantes_semanticas", "semantic_variants", "sinonimos", "sinónimos"]:
            col_map[c] = "variantes_semanticas"
        elif nc in ["required_items", "required_number_of_items", "numero_requerido", "n_requerido"]:
            col_map[c] = "required_items"
        elif nc in ["prompt", "enunciado", "pregunta_texto", "instruccion", "instrucción", "consigna"]:
            col_map[c] = "prompt"

    df = df.rename(columns=col_map)

    if "pregunta" not in df.columns:
        raise ValueError("La rúbrica Excel debe incluir una columna llamada 'pregunta'.")

    questions = []
    total_score = 0.0

    for _, row in df.iterrows():
        if pd.isna(row.get("pregunta", "")) or not str(row.get("pregunta", "")).strip():
            continue
        q = build_question_from_excel_row(row)
        questions.append(q)
        total_score += float(q.get("max_score", 0))

    if not questions:
        raise ValueError("La rúbrica Excel no contiene preguntas válidas.")

    required_columns = ["student_id", "nombre"] + [q["id"] for q in questions]

    rubric = {
        "name": path.stem.replace("_", " ").title(),
        "total_score": total_score,
        "input_format": {
            "required_columns": required_columns
        },
        "questions": questions,
        "_rubric_filename": path.name,
        "_rubric_display_name": path.stem.replace("_", " ").title()
    }

    return rubric


async def load_uploaded_rubric(rubric_file: UploadFile):
    suffix = Path(rubric_file.filename).suffix.lower()
    original_name = Path(rubric_file.filename).name
    temp_path = OUTPUT_DIR / f"uploaded_rubric_{original_name}"

    with open(temp_path, "wb") as f:
        f.write(await rubric_file.read())

    if suffix == ".json":
        rubric = load_rubric_from_json_path(temp_path)
        rubric["_rubric_filename"] = original_name
        rubric["_rubric_display_name"] = rubric.get("name") or rubric.get("title") or Path(original_name).stem.replace("_", " ").title()
        return rubric

    if suffix in [".xlsx", ".xls"]:
        rubric = load_rubric_from_excel(temp_path)
        rubric["_rubric_filename"] = original_name
        rubric["_rubric_display_name"] = Path(original_name).stem.replace("_", " ").title()
        return rubric

    raise ValueError("La rúbrica debe estar en formato .json, .xlsx o .xls.")


# ============================================================
# MOTOR CRB
# ============================================================



# ============================================================
# MOTOR SEMÁNTICO FINO v2.9
# ============================================================

STOPWORDS_ES = {
    "a","al","algo","ante","antes","como","con","contra","cual","cuando","de","del","desde",
    "donde","durante","e","el","ella","ellas","ellos","en","entre","era","eran","es","esa",
    "esas","ese","eso","esos","esta","estas","este","esto","estos","fue","fueron","ha","han",
    "hay","la","las","le","les","lo","los","mas","más","me","mi","mis","muy","no","nos",
    "o","para","pero","por","porque","que","qué","se","ser","si","sí","sin","sobre","su",
    "sus","tambien","también","te","tiene","tienen","un","una","unas","uno","unos","y","ya",
    "explica","explique","mencione","nombre","indique","complete","relacione","defina","respuesta",
    "pregunta","segun","según","respecto","acerca","sobre"
}

NEGATION_TERMS = {"no", "nunca", "jamás", "jamas", "sin", "tampoco", "niega", "negacion", "negación"}

# Banco conservador: no pretende cubrir todo el currículum; sirve para detectar errores fuertes.
CONCEPTUAL_CONTRADICTIONS = {
    "estrella": ["planeta", "satélite natural", "satelite natural", "luna"],
    "planeta": ["estrella", "galaxia"],
    "luna": ["planeta", "estrella"],
    "galaxia": ["planeta", "estrella individual"],
    "via lactea": ["andromeda"],
    "vía láctea": ["andromeda"],
    "fusión nuclear": ["combustión", "fuego", "quema carbón", "quema carbon"],
    "fusion nuclear": ["combustión", "fuego", "quema carbón", "quema carbon"],
    "gravedad": ["magnetismo solamente", "empuje", "viento"],
    "orbita": ["caída recta", "caida recta", "reposo"],
    "órbita": ["caída recta", "caida recta", "reposo"],
    "monarquia": ["democracia plena"],
    "monarquía": ["democracia plena"],
    "revolucion francesa": ["independencia de estados unidos"],
    "revolución francesa": ["independencia de estados unidos"],
    "independencia": ["dependencia total"],
}

RELATION_PATTERNS = {
    "causalidad": ["causa", "provoca", "produce", "genera", "permite", "hace que", "debido a", "gracias a", "por eso", "por lo tanto", "por ello"],
    "mantencion": ["mantiene", "sostiene", "conserva", "retiene", "atrae", "mantener", "manteni"],
    "pertenencia": ["pertenece", "forma parte", "es parte", "incluye", "contiene", "compone"],
    "definicion": ["es", "son", "se define", "consiste", "corresponde", "se trata"],
    "funcion": ["sirve", "funciona", "participa", "cumple", "se encarga", "permite"],
    "comparacion": ["a diferencia", "similar", "distinto", "en cambio", "mientras que"]
}

CENTRALITY_MARKERS = ["principal", "fundamental", "central", "clave", "importante", "esencial", "causa", "produce", "genera"]

SEMANTIC_SYNONYMS = {
    "lenguaje": ["habla", "comunicación", "comunicacion", "expresión", "expresion", "idioma"],
    "mente": ["procesos mentales", "cognición", "cognicion", "pensamiento", "actividad mental"],
    "cognicion": ["cognición", "mente", "pensamiento", "procesos mentales"],
    "cognición": ["cognicion", "mente", "pensamiento", "procesos mentales"],
    "comprension": ["comprensión", "entendimiento", "interpretación", "interpretacion", "entender", "comprender"],
    "comprensión": ["comprension", "entendimiento", "interpretación", "interpretacion", "entender", "comprender"],
    "memoria": ["recuerdo", "almacenamiento", "retención", "retencion"],
    "atencion": ["atención", "concentración", "concentracion", "foco"],
    "atención": ["atencion", "concentración", "concentracion", "foco"],
    "inferencia": ["deducción", "deduccion", "interpretación", "interpretacion", "concluir", "conclusión"],
    "coherencia": ["sentido global", "continuidad", "organización de ideas", "organizacion de ideas", "relación entre ideas", "relacion entre ideas"],
    "cohesion": ["cohesión", "conexión textual", "conexion textual", "marcadores", "conectores"],
    "cohesión": ["cohesion", "conexión textual", "conexion textual", "marcadores", "conectores"],
    "significado": ["sentido", "contenido", "idea"],
    "discurso": ["texto", "enunciado", "producción", "produccion", "mensaje"],
    "gravedad": ["atracción gravitacional", "atraccion gravitacional", "fuerza gravitatoria", "fuerza de gravedad", "atracción", "atraccion"],
    "orbita": ["órbita", "movimiento orbital", "trayectoria orbital", "gira alrededor", "giran alrededor", "orbitar"],
    "órbita": ["orbita", "movimiento orbital", "trayectoria orbital", "gira alrededor", "giran alrededor", "orbitar"],
    "sol": ["estrella central", "astro central"],
    "planetas": ["cuerpos celestes", "mundos", "astros"],
    "estrella": ["astro luminoso", "cuerpo luminoso", "astro que emite luz", "cuerpo celeste luminoso"],
    "fusion nuclear": ["fusión nuclear", "reacciones nucleares", "reacción nuclear"],
    "fusión nuclear": ["fusion nuclear", "reacciones nucleares", "reacción nuclear"],
    "energia": ["energía", "emisión de energía", "potencia"],
    "energía": ["energia", "emisión de energía", "potencia"],
    "luz": ["luminosidad", "brillo", "emite luz", "ilumina"],
    "calor": ["temperatura", "energía térmica", "energia termica"],
    "revolucion": ["revolución", "cambio político", "cambio politico", "levantamiento", "transformación política"],
    "revolución": ["revolucion", "cambio político", "cambio politico", "levantamiento", "transformación política"],
    "monarquia": ["monarquía", "rey", "corona", "régimen monárquico", "regimen monarquico"],
    "monarquía": ["monarquia", "rey", "corona", "régimen monárquico", "regimen monarquico"],
    "igualdad": ["equidad", "mismos derechos", "derechos iguales"],
    "libertad": ["autonomía", "autonomia", "independencia", "derechos"],
    "fraternidad": ["hermandad", "solidaridad"],
    "ilustracion": ["ilustración", "ideas ilustradas", "razón", "razon", "pensamiento ilustrado"],
    "ilustración": ["ilustracion", "ideas ilustradas", "razón", "razon", "pensamiento ilustrado"],
    "pueblo": ["ciudadanía", "ciudadania", "personas", "clases populares", "sociedad"],
    "independencia": ["emancipación", "emancipacion", "liberación", "liberacion", "autonomía", "autonomia"],
    "colonias": ["territorios coloniales", "colonos", "poblaciones coloniales"],
    "estados unidos": ["eeuu", "usa", "norteamérica", "norteamerica"],
}

# Ampliación semántica v3.0: dominios escolares frecuentes y oceanografía.
SEMANTIC_SYNONYMS.update({
    "oceanos": ["océanos", "mares", "mar", "agua marina", "masas de agua"],
    "océanos": ["oceanos", "mares", "mar", "agua marina", "masas de agua"],
    "regulacion climatica": ["regulación climática", "regulacion del clima", "regular el clima", "regulación térmica", "temperatura global"],
    "regulación climática": ["regulacion climatica", "regulacion del clima", "regular el clima", "regulación térmica", "temperatura global"],
    "corrientes marinas": ["corrientes", "circulación oceánica", "circulacion oceanica", "movimiento del agua"],
    "absorcion de calor": ["absorción de calor", "almacena calor", "retiene calor", "absorbe energía", "absorbe energia"],
    "salinidad": ["sales", "cantidad de sal", "concentración de sal", "concentracion de sal"],
    "densidad": ["masa por volumen", "agua más densa", "agua mas densa"],
    "fitoplancton": ["plancton vegetal", "microalgas", "base de la cadena trófica", "base de la cadena trofica"],
    "presion": ["presión", "alta presión", "presion del agua"],
    "presión": ["presion", "alta presión", "presion del agua"],
    "profundidad": ["zona profunda", "océano profundo", "oceano profundo", "más profundo", "mas profundo"],
    "luz solar": ["luz", "luminosidad", "radiación solar", "radiacion solar"],
    "contaminacion": ["contaminación", "desechos", "basura", "residuos", "microplásticos", "microplasticos", "plásticos", "plasticos"],
    "contaminación": ["contaminacion", "desechos", "basura", "residuos", "microplásticos", "microplasticos", "plásticos", "plasticos"],
    "ecosistemas marinos": ["vida marina", "ambientes marinos", "hábitats marinos", "habitats marinos"],
    "biodiversidad": ["diversidad biológica", "diversidad biologica", "variedad de especies", "especies marinas"],
})


def semantic_tokens(text):
    txt = normalize_text(text)
    return [t for t in re.split(r"\s+", txt) if t and t not in STOPWORDS_ES and len(t) > 2]


def simple_stem(token):
    t = normalize_text(token)
    for suffix in ["aciones", "iciones", "mente", "ación", "acion", "idades", "idad", "amiento", "imiento", "ismos", "istas", "icos", "icas", "ados", "adas", "ido", "ida", "ción", "sión", "cion", "sion", "es", "s"]:
        if len(t) > len(suffix) + 4 and t.endswith(suffix):
            return t[:-len(suffix)]
    return t


def token_overlap_score(answer, target):
    a_tokens = set(simple_stem(t) for t in semantic_tokens(answer))
    t_tokens = set(simple_stem(t) for t in semantic_tokens(target))
    if not a_tokens or not t_tokens:
        return 0
    overlap = len(a_tokens.intersection(t_tokens)) / len(t_tokens)
    containment = len(a_tokens.intersection(t_tokens)) / max(min(len(a_tokens), len(t_tokens)), 1)
    return int(round((0.70 * overlap + 0.30 * containment) * 100))


def answer_length_profile(answer):
    n = len(semantic_tokens(answer))
    if n <= 0:
        return "empty"
    if n <= 3:
        return "brief"
    if n <= 8:
        return "short"
    return "developed"


def synonym_expansions(term):
    norm = normalize_text(term)
    expansions = {term, norm}
    if norm in SEMANTIC_SYNONYMS:
        expansions.update(SEMANTIC_SYNONYMS[norm])
    for key, values in SEMANTIC_SYNONYMS.items():
        norm_values = [normalize_text(v) for v in values]
        if norm == normalize_text(key) or norm in norm_values:
            expansions.add(key)
            expansions.update(values)
    for tok in semantic_tokens(norm):
        if tok in SEMANTIC_SYNONYMS:
            expansions.update(SEMANTIC_SYNONYMS[tok])
    return [e for e in expansions if e]


def has_negation_near(answer, target, window=4):
    a_tokens = normalize_text(answer).split()
    target_stems = set(simple_stem(t) for t in semantic_tokens(target))
    if not a_tokens or not target_stems:
        return False
    stems = [simple_stem(t) for t in a_tokens]
    neg_positions = [i for i, t in enumerate(stems) if t in NEGATION_TERMS or normalize_text(a_tokens[i]) in NEGATION_TERMS]
    for i, stem in enumerate(stems):
        if stem in target_stems:
            for npos in neg_positions:
                if 0 <= i - npos <= window:
                    return True
    return False


def semantic_match_basic(answer, target):
    answer_n = normalize_text(answer)
    target_n = normalize_text(target)
    if not answer_n or not target_n:
        return False, 0, "vacío"
    if target_n in answer_n:
        return True, 100, "coincidencia directa"
    score = fuzz.partial_ratio(answer_n, target_n)
    if score >= 78:
        return True, int(score), "fuzzy"
    overlap = token_overlap_score(answer_n, target_n)
    if overlap >= 64:
        return True, int(overlap), "solapamiento conceptual"
    return False, int(max(score, overlap)), "sin coincidencia"


def semantic_match_uncached(answer, target, threshold=68, semantic_threshold=62):
    answer_n = normalize_text(answer)
    target_n = normalize_text(target)
    if not answer_n or not target_n:
        return False, 0, "vacío"

    if has_negation_near(answer_n, target_n):
        return False, 20, "posible negación o contradicción"

    # Caso breve válido: si la respuesta del estudiante contiene exactamente el núcleo conceptual,
    # no se le exige desarrollo textual adicional.
    if target_n in answer_n or answer_n in target_n:
        score = 100 if target_n in answer_n else 86
        return True, score, "núcleo conceptual breve"

    best_score = fuzz.partial_ratio(answer_n, target_n)
    best_method = "fuzzy"
    if best_score >= threshold:
        return True, int(best_score), best_method

    for expansion in synonym_expansions(target_n):
        exp_n = normalize_text(expansion)
        if not exp_n:
            continue
        if exp_n in answer_n or answer_n in exp_n:
            return True, 88, "sinónimo/paráfrasis breve"
        s = fuzz.partial_ratio(answer_n, exp_n)
        if s > best_score:
            best_score = s
            best_method = "sinónimo/paráfrasis"
        if s >= semantic_threshold + 10:
            return True, int(s), "sinónimo/paráfrasis"

    overlap = token_overlap_score(answer_n, target_n)
    if overlap > best_score:
        best_score = overlap
        best_method = "solapamiento conceptual"
    if overlap >= semantic_threshold:
        return True, int(overlap), "solapamiento conceptual"

    # Capa v3.3: embeddings semánticos reales cuando están disponibles.
    # Se usa como apoyo de recuperación conceptual, especialmente en respuestas breves
    # que no comparten suficientes palabras con la rúbrica.
    emb = embedding_similarity_score(answer_n, target_n)
    if emb is not None:
        emb_score, emb_cosine = emb
        if emb_score > best_score:
            best_score = emb_score
            best_method = f"embedding semántico cos={emb_cosine}"
        # Umbrales conservadores para evitar falsos positivos.
        if emb_score >= 82:
            return True, int(emb_score), f"embedding semántico cos={emb_cosine}"
        if answer_length_profile(answer_n) in ["brief", "short"] and emb_score >= 78:
            return True, int(emb_score), f"embedding semántico breve cos={emb_cosine}"

    return False, int(best_score), best_method


def semantic_match(answer, target, threshold=68, semantic_threshold=62):
    key = cache_key("semantic_match", answer, target, threshold, semantic_threshold)
    if key in SEMANTIC_CACHE:
        return SEMANTIC_CACHE[key]
    result = semantic_match_uncached(answer, target, threshold=threshold, semantic_threshold=semantic_threshold)
    if len(SEMANTIC_CACHE) < 20000:
        SEMANTIC_CACHE[key] = result
    return result


def contradictions_lookup(term):
    term_n = normalize_text(term)
    bad = []
    for key, values in CONCEPTUAL_CONTRADICTIONS.items():
        key_n = normalize_text(key)
        if term_n == key_n or term_n in [normalize_text(v) for v in synonym_expansions(key_n)]:
            bad.extend(values)
    return bad


def explicit_wrong_relation(answer, concept, bad):
    """
    Detecta error conceptual contextual.
    Penaliza definiciones o equivalencias incorrectas; evita penalizar negaciones correctas.
    """
    answer_n = normalize_text(answer)
    concept_n = normalize_text(concept)
    bad_n = normalize_text(bad)

    if not answer_n or not concept_n or not bad_n:
        return False

    negated_patterns = [
        f"{concept_n} no es {bad_n}",
        f"{concept_n} no son {bad_n}",
        f"{concept_n} no corresponde a {bad_n}",
        f"{concept_n} no funciona como {bad_n}",
        f"no es {bad_n}",
        f"no son {bad_n}",
    ]
    if any(p in answer_n for p in negated_patterns):
        return False

    wrong_patterns = [
        f"{concept_n} es {bad_n}",
        f"{concept_n} son {bad_n}",
        f"{concept_n} corresponde a {bad_n}",
        f"{concept_n} corresponde al {bad_n}",
        f"{concept_n} funciona como {bad_n}",
        f"{concept_n} se define como {bad_n}",
        f"{concept_n} consiste en {bad_n}",
        f"es un {bad_n}",
        f"es una {bad_n}",
        f"son {bad_n}",
    ]
    if any(p in answer_n for p in wrong_patterns):
        return True

    tokens = answer_n.split()
    concept_tokens = concept_n.split()
    bad_tokens = bad_n.split()

    def find_positions(seq):
        pos = []
        if not seq:
            return pos
        for i in range(len(tokens)):
            if " ".join(tokens[i:i+len(seq)]) == " ".join(seq):
                pos.append(i)
        return pos

    cpos = find_positions(concept_tokens)
    bpos = find_positions(bad_tokens)
    definitional = any(v in f" {answer_n} " for v in [" es ", " son ", " corresponde ", " funciona como ", " se define ", " consiste "])

    if definitional:
        for c in cpos:
            for b in bpos:
                if abs(c - b) <= 7:
                    return True
    return False


def detect_contradictions(answer, expected_terms):
    answer_n = normalize_text(answer)
    if not answer_n:
        return []

    found = []
    all_expected = []
    for term in expected_terms:
        all_expected.append(term)
        all_expected.extend(synonym_expansions(term))

    for term in all_expected:
        t = normalize_text(term)
        if not t:
            continue
        concept_present = t in answer_n or fuzz.partial_ratio(answer_n, t) >= 78
        if not concept_present:
            continue
        for bad in contradictions_lookup(t):
            bad_n = normalize_text(bad)
            if bad_n and (bad_n in answer_n or fuzz.partial_ratio(answer_n, bad_n) >= 88):
                if explicit_wrong_relation(answer_n, t, bad_n):
                    found.append(f"{term} ↔ {bad}")
    return sorted(set(found))


def detect_present_concepts(answer, concepts):
    present = []
    evidence = {}
    for c in concepts:
        best_score = 0
        best_method = "sin coincidencia"
        hit = False
        variants = [c] + synonym_expansions(c)
        seen = set()
        variants = [v for v in variants if not (normalize_text(v) in seen or seen.add(normalize_text(v)))]
        for v in variants:
            ok, score, method = semantic_match(answer, v, threshold=68, semantic_threshold=58)
            if score > best_score:
                best_score = score
                best_method = method
            if ok:
                hit = True
        if hit:
            present.append(c)
        evidence[c] = {"hit": hit, "score": best_score, "method": best_method}
    return sorted(set(present)), evidence


def relation_score(answer, concepts):
    answer_n = normalize_text(answer)
    if not answer_n or len(concepts) < 2:
        return 0, [], []

    present, _evidence = detect_present_concepts(answer_n, concepts)
    if len(present) < 2:
        return 0, [], present

    relation_hits = []
    for rel, patterns in RELATION_PATTERNS.items():
        for pat in patterns:
            pat_n = normalize_text(pat)
            if pat_n and pat_n in answer_n:
                relation_hits.append(rel)
                break

    if relation_hits:
        return min(0.20, 0.07 * len(set(relation_hits))), sorted(set(relation_hits)), present

    # Si hay dos o más conceptos centrales en una respuesta desarrollada, se reconoce integración mínima,
    # pero con bono menor que una relación explícita.
    if len(answer_n.split()) >= 8:
        return 0.05, ["integración conceptual básica"], present
    return 0, [], present


def infer_concept_weights(concepts, prompt=""):
    concepts = [c for c in concepts if str(c).strip()]
    if not concepts:
        return {}
    prompt_n = normalize_text(prompt)
    raw = []
    for i, c in enumerate(concepts):
        c_n = normalize_text(c)
        w = 1.0
        if i == 0:
            w += 0.25
        if c_n and c_n in prompt_n:
            w += 0.18
        if any(m in c_n for m in CENTRALITY_MARKERS):
            w += 0.25
        if len(c_n.split()) >= 2:
            w += 0.08
        raw.append(w)
    total = sum(raw) or 1
    return {concepts[i]: raw[i] / total for i in range(len(concepts))}


def fuzzy_contains(answer, target, threshold=78):
    ok, score, _method = semantic_match(answer, target, threshold=threshold, semantic_threshold=max(58, threshold - 10))
    return ok, score


VAGUE_RESPONSES = {
    "no se", "no sé", "nose", "porque si", "porque sí", "no entiendo", "no recuerdo",
    "no tengo idea", "nada", "ninguna", "sin respuesta", "no aplica"
}


def get_question_concepts(question):
    item_type = question.get("item_type", "")
    if item_type == "criteria":
        return [c.get("concept", "") for c in question.get("criteria", []) if c.get("concept", "")]
    if item_type in ["enumeration", "enumeration_closed", "enumeration_conceptual", "enumeration_categorized"]:
        return question.get("accepted_concepts", question.get("accepted_answers", []))
    if item_type in ["completion", "short_exact_answer", "true_false"]:
        return question.get("accepted_answers", [])
    if item_type == "classification_matching":
        concepts = []
        for pair in question.get("pairs", []):
            concepts.append(pair.get("prompt_value", ""))
            concepts.append(pair.get("correct_match", ""))
        return [c for c in concepts if c]
    return []


CONCEPT_RELATION_MARKERS_V31 = {
    "causalidad": ["causa", "provoca", "produce", "genera", "debido a", "por eso", "gracias a"],
    "funcion": ["sirve", "permite", "ayuda", "cumple", "se encarga", "participa"],
    "definicion": ["es", "son", "se define", "corresponde", "consiste"],
    "consecuencia": ["afecta", "daña", "impacta", "consecuencia", "resultado"],
    "oposicion": ["pero", "sin embargo", "en cambio", "a diferencia"],
    "comparacion": ["más que", "menos que", "mayor", "menor", "similar", "distinto"],
    "gradiente": ["aumenta", "disminuye", "sube", "baja", "profundidad", "temperatura"],
}

def detect_concept_relations(answer, concepts):
    answer_n = normalize_text(answer)
    present, _evidence = detect_present_concepts(answer, concepts)
    relation_hits = []
    for rel, patterns in RELATION_PATTERNS.items():
        for pat in patterns:
            if normalize_text(pat) in answer_n:
                relation_hits.append(rel)
                break
    # Relaciones escolares frecuentes y mapa ampliado v3.1.
    extra = dict(CONCEPT_RELATION_MARKERS_V31)
    extra.update({
        "cambio_gradual": ["aumenta", "disminuye", "más", "menos", "mayor", "menor", "cambia"],
        "regulacion": ["regula", "modera", "mantiene", "estabiliza"],
    })
    for rel, patterns in extra.items():
        for pat in patterns:
            if normalize_text(pat) in answer_n:
                relation_hits.append(rel)
                break
    return sorted(set(relation_hits)), present


def classify_error_type(answer, question, concepts=None, present=None, missing=None, contradictions=None):
    answer_n = normalize_text(answer)
    if not answer_n:
        return "respuesta_vacia", "alta"
    if answer_n in {normalize_text(x) for x in VAGUE_RESPONSES} or len(semantic_tokens(answer_n)) <= 1:
        return "respuesta_vaga", "media"

    concepts = concepts if concepts is not None else get_question_concepts(question)
    present = present if present is not None else detect_present_concepts(answer, concepts)[0]
    missing = missing if missing is not None else [c for c in concepts if c not in present]
    contradictions = contradictions if contradictions is not None else detect_contradictions(answer, concepts)

    if contradictions:
        return "confusion_o_contradiccion_conceptual", "alta"
    if concepts and not present:
        return "fuera_de_foco_u_omision_total", "alta"
    if missing and present:
        if answer_length_profile(answer) in ["brief", "short"]:
            return "respuesta_breve_parcial", "baja"
        return "respuesta_incompleta", "media"

    relations, _ = detect_concept_relations(answer, concepts)
    if len(concepts) >= 2 and len(present) >= 2 and not relations and answer_length_profile(answer) == "developed":
        return "relacion_conceptual_debil", "media"

    return "sin_error_conceptual_evidente", "baja"


def generate_teacher_feedback_from_diagnosis(error_type, missing, relations, answer_profile):
    if error_type == "respuesta_vacia":
        return "Solicitar respuesta; no hay evidencia evaluable."
    if error_type == "respuesta_vaga":
        return "Pedir mayor precisión conceptual; la respuesta es demasiado general."
    if error_type == "confusion_o_contradiccion_conceptual":
        return "Revisar manualmente: puede existir una confusión conceptual relevante."
    if error_type == "fuera_de_foco_u_omision_total":
        return "Reforzar el núcleo conceptual de la pregunta; no se detectan ideas esperadas."
    if error_type == "respuesta_breve_parcial":
        return "La respuesta contiene un núcleo válido, pero conviene pedir desarrollo si la rúbrica exige explicación."
    if error_type == "respuesta_incompleta":
        return "Retroalimentar conceptos omitidos: " + (", ".join(missing[:4]) if missing else "sin detalle") + "."
    if error_type == "relacion_conceptual_debil":
        return "Pedir que conecte explícitamente los conceptos mediante causa, función o consecuencia."
    if relations:
        return "Respuesta conceptualmente consistente; se detectan relaciones: " + ", ".join(relations[:4]) + "."
    if answer_profile in ["brief", "short"]:
        return "Respuesta aceptable pero breve; revisar si el objetivo era explicación desarrollada."
    return "Respuesta compatible con los criterios actuales de Evalia."


def semantic_diagnosis(answer, question, score=None, confidence=None, status=None):
    concepts = get_question_concepts(question)
    present, evidence = detect_present_concepts(answer, concepts)
    missing = [c for c in concepts if c not in present]
    contradictions = detect_contradictions(answer, concepts)
    relations, present_for_relations = detect_concept_relations(answer, concepts)
    profile = answer_length_profile(answer)
    error_type, severity = classify_error_type(
        answer, question, concepts=concepts, present=present, missing=missing, contradictions=contradictions
    )
    coverage = len(set(present)) / max(len(set(concepts)), 1) if concepts else 0
    teacher_suggestion = generate_teacher_feedback_from_diagnosis(error_type, missing, relations, profile)
    return {
        "answer_profile": profile,
        "concepts_expected": "; ".join([str(c) for c in concepts]),
        "concepts_detected": "; ".join([str(c) for c in present]),
        "concepts_missing": "; ".join([str(c) for c in missing]),
        "conceptual_relations": "; ".join(relations),
        "contradictions": "; ".join(contradictions),
        "error_type": error_type,
        "error_severity": severity,
        "conceptual_coverage": round(coverage, 2),
        "teacher_suggestion_semantic": teacher_suggestion,
        "embedding_mode": EMBEDDING_STATUS.get("mode", "reglas_semanticas"),
        "evalia_decision_basis": f"cobertura={coverage:.2f}; relaciones={len(relations)}; contradicciones={len(contradictions)}; perfil={profile}; embeddings={EMBEDDING_STATUS.get('mode', 'reglas_semanticas')}"
    }


def score_accepted_answers(answer, question):
    accepted = question.get("accepted_answers", [])
    if not accepted:
        return 0, 0.0, "Sin respuestas aceptadas configuradas."

    best = 0
    best_method = "sin coincidencia"
    best_target = ""
    for target in accepted:
        ok, score, method = semantic_match(answer, target, threshold=80, semantic_threshold=66)
        if score > best:
            best = score
            best_method = method
            best_target = target
        if ok:
            return question["max_score"], min(score / 100, 1.0), f"Respuesta compatible con: {target} ({method})."
    return 0, best / 100, f"No coincide suficientemente con las respuestas aceptadas. Mejor aproximación: {best_target} ({best_method}, {best}/100)."


def score_true_false(answer, question):
    a = normalize_text(answer)
    accepted = [normalize_text(x) for x in question.get("accepted_answers", [])]
    if a in accepted:
        return question["max_score"], 1.0, "Respuesta cerrada correcta."
    if accepted and a[:1] == accepted[0][:1]:
        return question["max_score"], 0.9, "Respuesta cerrada correcta por inicial."
    return 0, 0.9 if a else 0.2, "Respuesta cerrada incorrecta o vacía."


def score_criteria(answer, question):
    criteria = question.get("criteria", [])
    if not criteria:
        return score_accepted_answers(answer, question)

    max_score = float(question.get("max_score", 1.0))
    concepts = [c.get("concept", "") for c in criteria if c.get("concept", "")]
    weight_map = infer_concept_weights(concepts, question.get("prompt", ""))

    total = 0.0
    matched = []
    missing = []
    confidence_scores = []
    method_notes = []

    contradictions = detect_contradictions(answer, concepts)
    profile = answer_length_profile(answer)

    for criterion in criteria:
        concept = criterion.get("concept", "")
        if not concept:
            continue

        base_weight = float(criterion.get("weight", 1.0))
        auto_weight = weight_map.get(concept)
        weight = max_score * auto_weight if auto_weight is not None and len(criteria) > 1 else base_weight

        variants = [concept]
        variants.extend(criterion.get("semantic_variants", []))
        variants.extend(criterion.get("accepted_values", []))
        variants.extend(synonym_expansions(concept))

        seen = set()
        variants = [v for v in variants if v and not (normalize_text(v) in seen or seen.add(normalize_text(v)))]

        criterion_best = 0
        criterion_method = "sin coincidencia"
        criterion_hit = False
        for v in variants:
            ok, score, method = semantic_match(answer, v, threshold=68, semantic_threshold=58)
            if score > criterion_best:
                criterion_best = score
                criterion_method = method
            if ok:
                criterion_hit = True

        related_contradiction = any(normalize_text(concept) in normalize_text(c) for c in contradictions)
        if related_contradiction:
            criterion_best = min(criterion_best, 42)
            criterion_hit = False
            criterion_method = "error conceptual probable"

        # Respuesta breve válida: reconoce el núcleo conceptual sin exigir desarrollo cuando el ítem tiene varios criterios.
        if criterion_hit and profile in ["brief", "short"]:
            confidence_scores.append(max(0.72, criterion_best / 100))
        else:
            confidence_scores.append(criterion_best / 100 if criterion_best else 0)

        if criterion_hit:
            total += weight
            matched.append(concept)
            method_notes.append(f"{concept}: {criterion_method} ({criterion_best}/100)")
        else:
            missing.append(concept)
            if criterion_best >= 40:
                method_notes.append(f"{concept}: aproximación insuficiente ({criterion_best}/100)")

    rel_bonus, rel_hits, present_for_relation = relation_score(answer, concepts)
    if rel_bonus and matched:
        total += max_score * rel_bonus
        method_notes.append("Relaciones detectadas: " + ", ".join(rel_hits))

    coverage = len(set(matched)) / max(len(set(concepts)), 1)

    # Ajuste epistemológico: una respuesta breve puede ser correcta, pero no siempre completa.
    # Si hay un solo criterio y coincide, puede recibir el total. Si hay varios criterios, recibe lo que cubre.
    if profile == "brief" and matched and not contradictions:
        method_notes.append("Respuesta breve válida: se detecta núcleo conceptual, aunque con desarrollo limitado.")

    if contradictions:
        contradiction_penalty = min(0.38, 0.16 + 0.08 * (len(contradictions) - 1))
        total *= (1 - contradiction_penalty)
        method_notes.append("Alerta: posible error conceptual: " + "; ".join(contradictions[:4]))

    total = min(max(total, 0), max_score)

    avg_evidence = sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0.0
    confidence = (0.55 * avg_evidence) + (0.30 * coverage) + (0.15 * (1 if not contradictions else 0.35))

    # Si hay respuesta breve correcta, subir confianza operacional, pero no fabricar completitud.
    if profile in ["brief", "short"] and matched and not contradictions:
        confidence = max(confidence, 0.68 if coverage < 0.80 else 0.82)

    if contradictions:
        confidence = min(confidence, 0.58)

    feedback = []
    if matched:
        feedback.append("Criterios detectados: " + "; ".join([m for m in matched if m]))
    if missing:
        feedback.append("Criterios no detectados o débiles: " + "; ".join([m for m in missing if m]))
    if profile in ["brief", "short"] and matched:
        feedback.append("Observación: respuesta breve; Evalia reconoce el núcleo, pero sugiere revisar si el docente esperaba desarrollo argumentativo.")
    if method_notes:
        feedback.append("Evidencia semántica: " + "; ".join(method_notes[:12]))

    return round(total, 2), round(confidence, 2), " | ".join(feedback)


def score_enumeration(answer, question):
    max_score = float(question.get("max_score", 0))
    required = question.get("constraints", {}).get("required_number_of_items")
    accepted = question.get("accepted_concepts", question.get("accepted_answers", []))

    hits = []
    notes = []
    for concept in accepted:
        variants = [concept] + synonym_expansions(concept)
        best_score = 0
        best_method = "sin coincidencia"
        hit = False
        for v in variants:
            ok, score, method = semantic_match(answer, v, threshold=68, semantic_threshold=58)
            if score > best_score:
                best_score = score
                best_method = method
            if ok:
                hit = True
        if hit:
            hits.append(concept)
            notes.append(f"{concept}: {best_method}")

    if required is None:
        required = len(accepted) if accepted else 1
    counted = min(len(set(hits)), required)
    score = max_score * (counted / required) if required else 0
    confidence = counted / required if required else 0
    detail = f"Elementos válidos detectados: {counted}/{required}. {', '.join(sorted(set(hits)))}"
    if notes:
        detail += " | Evidencia semántica: " + "; ".join(notes[:8])
    return round(score, 2), round(confidence, 2), detail


def score_matching(answer, question):
    max_score = float(question.get("max_score", 0))
    pairs = question.get("pairs", [])
    total = 0.0
    found = []

    for pair in pairs:
        left = pair.get("prompt_value", "")
        right = pair.get("correct_match", "")
        weight = float(pair.get("weight", 1.0))
        left_variants = [left] + synonym_expansions(left)
        right_variants = [right] + synonym_expansions(right)
        ok_left = False
        ok_right = False
        method_left = ""
        method_right = ""
        for v in left_variants:
            ok, score, method = semantic_match(answer, v, threshold=65, semantic_threshold=58)
            if ok:
                ok_left = True
                method_left = method
                break
        for v in right_variants:
            ok, score, method = semantic_match(answer, v, threshold=65, semantic_threshold=58)
            if ok:
                ok_right = True
                method_right = method
                break
        if ok_left and ok_right:
            total += weight
            found.append(f"{left} -> {right} ({method_left}/{method_right})")
    confidence = total / max_score if max_score else 0
    return round(min(total, max_score), 2), round(confidence, 2), "Relaciones detectadas: " + "; ".join(found)


def score_answer(answer, question):
    item_type = question.get("item_type", "")

    if not str(answer).strip():
        return 0, 0.1, "Respuesta vacía.", "revisar"

    if item_type == "true_false":
        score, conf, fb = score_true_false(answer, question)
    elif item_type in ["completion", "short_exact_answer"]:
        score, conf, fb = score_accepted_answers(answer, question)
    elif item_type in ["enumeration", "enumeration_closed", "enumeration_conceptual", "enumeration_categorized"]:
        score, conf, fb = score_enumeration(answer, question)
    elif item_type == "classification_matching":
        score, conf, fb = score_matching(answer, question)
    else:
        score, conf, fb = score_criteria(answer, question)

    # Tres estados estables para docentes: alta confianza, parcial/intermedia y revisión sugerida.
    if conf >= 0.82:
        status = "aceptado"
    elif conf < 0.48:
        status = "revisar"
    else:
        status = "aceptado_con_cautela"
    return score, conf, fb, status


def performance_level(pct):
    """Clasificación docente simple del porcentaje total o por ítem."""
    try:
        pct = float(pct)
    except Exception:
        return "Sin información"
    if pct >= 80:
        return "Alto"
    if pct >= 60:
        return "Medio"
    return "Bajo"



# ============================================================
# VALIDACIÓN PREVENTIVA v3.1
# ============================================================

def validate_rubric_integrity(rubric):
    issues = []
    questions = rubric.get("questions", []) if isinstance(rubric, dict) else []
    if not questions:
        issues.append("La rúbrica no contiene preguntas válidas.")
        return issues
    seen = set()
    for i, q in enumerate(questions, start=1):
        qid = str(q.get("id", "")).strip()
        if not qid:
            issues.append(f"La pregunta {i} no tiene identificador.")
        elif qid in seen:
            issues.append(f"El identificador de pregunta está duplicado: {qid}.")
        seen.add(qid)
        try:
            max_score = float(q.get("max_score", 0))
            if max_score <= 0:
                issues.append(f"{qid}: el puntaje debe ser mayor que 0.")
        except Exception:
            issues.append(f"{qid}: el puntaje no es numérico.")
        item_type = normalize_item_type(q.get("item_type", "criteria"))
        if item_type == "criteria" and not q.get("criteria") and not q.get("accepted_answers"):
            issues.append(f"{qid}: pregunta abierta sin criterios ni ideas esperadas.")
        if item_type in ["completion", "short_exact_answer", "true_false"] and not q.get("accepted_answers"):
            issues.append(f"{qid}: pregunta cerrada sin respuesta esperada.")
        if item_type == "classification_matching" and not q.get("pairs"):
            issues.append(f"{qid}: pregunta de relacionar sin pares válidos. Usa Concepto:Respuesta; Concepto:Respuesta.")
    return issues

def validate_dataframe_integrity(df):
    issues = []
    if df is None or df.empty:
        issues.append("El archivo de respuestas está vacío.")
        return issues
    if len(df.columns) < 3:
        issues.append("El archivo de respuestas debe incluir identificación, nombre y al menos una pregunta.")
    unnamed = [str(c) for c in df.columns if str(c).lower().startswith("unnamed")]
    if len(unnamed) >= max(2, len(df.columns)//2):
        issues.append("El Excel parece tener muchas columnas vacías o sin nombre. Revisa la fila de encabezados.")
    if len(df) > 5000:
        issues.append("El archivo contiene más de 5000 filas. Para esta versión se recomienda dividir el procesamiento.")
    return issues

def build_traceability_row(student_id, nombre, pregunta_id, answer, score, confidence, status, diagnosis, feedback):
    return {
        "student_id": student_id,
        "nombre": nombre,
        "pregunta_id": pregunta_id,
        "status_evalia": status,
        "score": score,
        "confidence": confidence,
        "perfil_respuesta": diagnosis.get("answer_profile", ""),
        "cobertura_conceptual": diagnosis.get("conceptual_coverage", ""),
        "relaciones_detectadas": diagnosis.get("conceptual_relations", ""),
        "tipo_error": diagnosis.get("error_type", ""),
        "severidad_error": diagnosis.get("error_severity", ""),
        "base_decision": diagnosis.get("evalia_decision_basis", ""),
        "feedback_evalia": feedback,
        "respuesta_original": answer
    }

# ============================================================
# VALIDACIÓN FLEXIBLE
# ============================================================

def validate_columns_flexible(df, rubric):
    """
    Valida con flexibilidad:
    - student_id puede ser id, alumno_id, rut, codigo...
    - nombre puede ser name, estudiante, alumno...
    - P1 puede calzar con Q1, pregunta1, item1...
    """
    missing = []

    if not find_column(df, "student_id", ["id", "alumno_id", "estudiante_id", "rut", "codigo", "código"]):
        missing.append("student_id/id")

    if not find_column(df, "nombre", ["name", "student_name", "alumno", "estudiante", "nombre_estudiante"]):
        missing.append("nombre/name")

    for q in rubric.get("questions", []):
        pid = q.get("id")
        if not find_item_column(df, pid):
            aliases = sorted(item_aliases(pid))
            missing.append(f"{pid} ({' / '.join(list(aliases)[:5])}...)")

    return missing


# ============================================================
# INSIGHTS
# ============================================================


def pedagogical_item_suggestion(classification, review_pct, avg_score_pct, avg_confidence):
    """Sugerencia docente simple y explicable por ítem."""
    classification = str(classification or "").lower()
    try:
        review_pct = float(review_pct)
        avg_score_pct = float(avg_score_pct)
        avg_confidence = float(avg_confidence)
    except Exception:
        return "Revisar manualmente una muestra de respuestas para confirmar el funcionamiento del ítem."

    if "revisión" in classification or "revision" in classification or "problem" in classification:
        return "Revisar el enunciado, la rúbrica o los criterios de corrección; varios estudiantes podrían haber interpretado la pregunta de manera distinta a lo esperado."
    if review_pct >= 30:
        return "Conviene revisar manualmente una muestra de respuestas antes de cerrar la calificación."
    if avg_score_pct < 60:
        return "La pregunta parece difícil para el grupo; puede requerir retroalimentación o refuerzo de contenidos."
    if avg_confidence < 0.70:
        return "La respuesta esperada podría necesitar criterios más explícitos o ejemplos adicionales."
    return "El ítem muestra funcionamiento estable bajo los criterios actuales de Evalia."


def build_question_insights(question_stats, questions):
    question_map = {q.get("id"): q for q in questions}
    insights_rows = []
    problematic_questions = []

    for pid, stats in question_stats.items():
        total = stats["total"] or 1
        max_score = float(question_map.get(pid, {}).get("max_score", 1)) or 1

        avg_score_raw = sum(stats["scores"]) / total if stats["scores"] else 0
        avg_score_pct = (avg_score_raw / max_score) * 100 if max_score else 0
        avg_confidence = sum(stats["confidences"]) / total if stats["confidences"] else 0

        accepted_pct = (stats["accepted"] / total) * 100
        caution_pct = (stats["caution"] / total) * 100
        review_pct = (stats["review"] / total) * 100

        if avg_score_pct >= 80 and review_pct < 20:
            classification = "funcionamiento alto"
        elif review_pct >= 35 or accepted_pct <= 50 or avg_confidence < 0.60:
            classification = "requiere revisión docente"
            problematic_questions.append(pid)
        elif caution_pct >= 30:
            classification = "funcionamiento parcial/intermedio"
        else:
            classification = "funcionamiento adecuado"

        insights_rows.append({
            "pregunta": pid,
            "tipo_item": display_item_type(question_map.get(pid, {}).get("item_type", "")),
            "puntaje_maximo": max_score,
            "promedio_puntaje": round(avg_score_raw, 2),
            "promedio_porcentaje": round(avg_score_pct, 1),
            "nivel_item": performance_level(avg_score_pct),
            "confianza_promedio": round(avg_confidence, 2),
            "aceptacion_pct": round(accepted_pct, 1),
            "cautela_pct": round(caution_pct, 1),
            "revision_pct": round(review_pct, 1),
            "clasificacion_evalia": classification,
            "sugerencia_docente": pedagogical_item_suggestion(classification, review_pct, avg_score_pct, avg_confidence)
        })

    return insights_rows, problematic_questions


def build_type_insights(insights_rows):
    if not insights_rows:
        return []

    df = pd.DataFrame(insights_rows)
    grouped = (
        df.groupby("tipo_item", dropna=False)
        .agg(
            preguntas=("pregunta", "count"),
            promedio_porcentaje=("promedio_porcentaje", "mean"),
            confianza_promedio=("confianza_promedio", "mean"),
            revision_pct=("revision_pct", "mean"),
            aceptacion_pct=("aceptacion_pct", "mean")
        )
        .reset_index()
    )

    rows = []
    for _, row in grouped.iterrows():
        rows.append({
            "tipo_item": row["tipo_item"] if row["tipo_item"] else "sin_tipo",
            "numero_preguntas": int(row["preguntas"]),
            "promedio_porcentaje": round(float(row["promedio_porcentaje"]), 1),
            "confianza_promedio": round(float(row["confianza_promedio"]), 2),
            "revision_promedio_pct": round(float(row["revision_pct"]), 1),
            "aceptacion_promedio_pct": round(float(row["aceptacion_pct"]), 1)
        })
    return rows


def build_interpretation(insights_rows, problematic_questions, total_students, rubric_name):
    if not insights_rows:
        return "No fue posible generar interpretación porque no se registraron métricas por pregunta."

    df = pd.DataFrame(insights_rows)
    avg_eval = float(df["promedio_porcentaje"].mean()) if "promedio_porcentaje" in df else 0
    avg_review = float(df["revision_pct"].mean()) if "revision_pct" in df else 0
    avg_conf = float(df["confianza_promedio"].mean()) if "confianza_promedio" in df else 0

    if avg_eval >= 80:
        level = "alto"
    elif avg_eval >= 60:
        level = "medio"
    else:
        level = "bajo"

    if problematic_questions:
        problem_text = (
            "Se identifican posibles focos de revisión en las preguntas "
            + ", ".join(problematic_questions)
            + ", debido a baja aceptación, baja confianza o alta proporción de respuestas marcadas para revisión."
        )
    else:
        problem_text = "No se detectan preguntas críticamente problemáticas bajo los criterios actuales de Evalia."

    return (
        f"La evaluación procesada con la rúbrica '{rubric_name}' incluyó {total_students} estudiante(s). "
        f"El desempeño global estimado por pregunta se ubica en un nivel {level}, "
        f"con un promedio general de {avg_eval:.1f}% del puntaje esperado. "
        f"La confianza promedio del sistema fue {avg_conf:.2f} y la tasa promedio de revisión manual fue {avg_review:.1f}%. "
        f"{problem_text} "
        "Estos resultados constituyen una primera capa de inteligencia evaluativa: permiten orientar la revisión docente, "
        "detectar ítems que podrían requerir ajuste y mejorar progresivamente la calidad de la evaluación."
    )



# ============================================================
# REPORTE EXCEL PREMIUM / DOCENTE
# ============================================================

def build_teacher_report_rows(score_rows, insights_rows, interpretation, problematic_questions, rubric_name, total_students, total_questions, auto_rate, caution_rate, review_rate):
    avg_pct = 0.0
    if score_rows:
        vals = [float(r.get("porcentaje", 0) or 0) for r in score_rows]
        avg_pct = sum(vals) / len(vals)

    if insights_rows:
        sorted_items = sorted(insights_rows, key=lambda x: float(x.get("promedio_porcentaje", 0) or 0))
        hardest = ", ".join([str(x.get("pregunta")) for x in sorted_items[:3]])
        best = ", ".join([str(x.get("pregunta")) for x in sorted_items[-3:][::-1]])
    else:
        hardest = "Sin información"
        best = "Sin información"

    return [
        {"seccion": "Síntesis", "indicador": "Rúbrica aplicada", "valor": rubric_name},
        {"seccion": "Síntesis", "indicador": "Estudiantes procesados", "valor": total_students},
        {"seccion": "Síntesis", "indicador": "Preguntas evaluadas", "valor": total_questions},
        {"seccion": "Resultados", "indicador": "Promedio general estimado", "valor": f"{avg_pct:.1f}%"},
        {"seccion": "Resultados", "indicador": "Nivel global", "valor": performance_level(avg_pct)},
        {"seccion": "Trazabilidad", "indicador": "Aceptación automática", "valor": f"{auto_rate}%"},
        {"seccion": "Trazabilidad", "indicador": "Aceptado con cautela", "valor": f"{caution_rate}%"},
        {"seccion": "Trazabilidad", "indicador": "Requiere revisión manual", "valor": f"{review_rate}%"},
        {"seccion": "Ítems", "indicador": "Preguntas con mayor dificultad", "valor": hardest},
        {"seccion": "Ítems", "indicador": "Preguntas con mejor funcionamiento", "valor": best},
        {"seccion": "Ítems", "indicador": "Preguntas potencialmente problemáticas", "valor": ", ".join(problematic_questions) if problematic_questions else "Sin preguntas críticas"},
        {"seccion": "Interpretación", "indicador": "Lectura automática", "valor": interpretation},
        {"seccion": "Sugerencia", "indicador": "Uso docente recomendado", "valor": "Usar este reporte como primera capa de revisión: confirmar manualmente los casos marcados como revisar y ajustar preguntas o criterios si un ítem aparece como problemático."},
    ]


def format_workbook(writer):
    wb = writer.book
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    soft_fill = PatternFill("solid", fgColor="F3F4F6")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    yellow_fill = PatternFill("solid", fgColor="FEF9C3")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    line = Side(style="thin", color="E5E7EB")
    border = Border(left=line, right=line, top=line, bottom=line)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if cell.row % 2 == 0:
                    cell.fill = soft_fill

        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

        def color_by_column(col_name):
            col_idx = headers.get(col_name)
            if not col_idx:
                return
            for r in range(2, ws.max_row + 1):
                value = ws.cell(r, col_idx).value
                txt = str(value).lower() if value is not None else ""
                fill = None
                if any(x in txt for x in ["alto", "aceptado", "funcionamiento alto", "estable"]):
                    fill = green_fill
                if any(x in txt for x in ["medio", "cautela"]):
                    fill = yellow_fill
                if any(x in txt for x in ["bajo", "revisar", "problemático", "problemat"]):
                    fill = red_fill
                if fill:
                    ws.cell(r, col_idx).fill = fill

        for cn in ["nivel_desempeno", "nivel_item", "status", "clasificacion_evalia"]:
            color_by_column(cn)

        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 70))
            ws.column_dimensions[letter].width = max(13, min(max_len + 2, 58))

        ws.auto_filter.ref = ws.dimensions

    if "REPORTE_DOCENTE" in wb.sheetnames:
        wb.move_sheet(wb["REPORTE_DOCENTE"], offset=-len(wb.sheetnames))
        ws = wb["REPORTE_DOCENTE"]
        ws.freeze_panes = "A2"
        for row in range(2, ws.max_row + 1):
            ws.cell(row, 1).font = Font(bold=True)
            ws.cell(row, 2).font = Font(bold=True)

# ============================================================
# CSS/UI
# ============================================================


# ============================================================
# CSS/UI
# ============================================================

def base_css():
    return """
    <style>
      :root {
        --bg: #f5f6f8; --card: #ffffff; --ink: #15171a; --muted: #667085;
        --line: #dbe4f0; --accent: #111827; --soft: #f8fbff; --ok: #067647;
        --alt-blue:#0ea5e9; --alt-purple:#7c3aed; --alt-orange:#f97316; --alt-green:#10b981;
        --evalia-soft:#eef7ff; --evalia-border:#b9ddff;
      }
      * { box-sizing: border-box; }
      body {
        margin: 0; min-height: 100vh;
        font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif;
        background: radial-gradient(circle at top left, rgba(14,165,233,.16), transparent 30%),
                    radial-gradient(circle at top right, rgba(124,58,237,.12), transparent 28%),
                    linear-gradient(180deg, #f7fbff 0%, #f5f6f8 55%, #f8fbff 100%);
        color: var(--ink);
      }
      .page { width: 100%; min-height: 100vh; display: flex; justify-content: center; padding: 42px 20px 24px; }
      .shell { width: 100%; max-width: 980px; }
      .topbar { display: flex; justify-content: space-between; align-items: center; margin-bottom: 26px; }
      .brand { display: flex; align-items: center; gap: 12px; }
      .logo {
        width: 44px; height: 44px; border-radius: 15px; background: #111827; color: white;
        display: grid; place-items: center; font-weight: 900; letter-spacing: -.03em;
        position: relative; overflow: hidden;
      }
      .logo:after {
        content:""; position:absolute; inset:auto 6px 6px 6px; height:4px; border-radius:999px;
        background: linear-gradient(90deg,var(--alt-blue),var(--alt-green),var(--alt-orange),var(--alt-purple));
      }
      .brand-title { font-size: 22px; font-weight: 900; letter-spacing: -.04em; }
      .brand-subtitle { color: var(--muted); font-size: 13px; margin-top: 2px; }
      .badge {
        padding: 8px 12px; background: white; border: 1px solid var(--line);
        border-radius: 999px; color: var(--muted); font-size: 13px;
      }
      .hero, .result-card {
        background: rgba(255,255,255,.90); backdrop-filter: blur(12px); border: 1px solid var(--line);
        border-radius: 28px; box-shadow: 0 20px 60px rgba(17,24,39,.08); overflow: hidden;
      }
      .hero-inner, .result-card { padding: 34px; }
      h1 { margin: 0; font-size: 42px; line-height: 1.04; letter-spacing: -.06em; }
      .lead { margin: 14px 0 26px; color: var(--muted); font-size: 17px; line-height: 1.55; max-width: 800px; }
      .panel { background: linear-gradient(180deg,#ffffff 0%, var(--evalia-soft) 100%); border: 1px solid var(--evalia-border); border-radius: 22px; padding: 22px; }
      .field-label { display: block; font-weight: 800; margin-bottom: 8px; font-size: 14px; }
      select, input.file-visible {
        width: 100%; padding: 14px; border: 1px solid var(--line); border-radius: 14px;
        background: white; font-size: 15px; color: var(--ink); margin-bottom: 18px;
      }
      .dropzone {
        position: relative; border: 2px dashed #9ecdfb; background: linear-gradient(180deg,#ffffff 0%,#f8fbff 100%); border-radius: 20px;
        padding: 28px 20px; text-align: center; transition: .18s ease; cursor: pointer; margin-bottom: 18px;
      }
      .dropzone:hover { border-color: #111827; transform: translateY(-1px); box-shadow: 0 12px 28px rgba(17,24,39,.08); }
      .dropzone input { position: absolute; inset: 0; opacity: 0; cursor: pointer; }
      .drop-icon { font-size: 30px; margin-bottom: 8px; }
      .drop-title { font-weight: 900; font-size: 17px; }
      .drop-subtitle { color: var(--muted); margin-top: 6px; font-size: 14px; }
      .file-name { margin-top: 10px; font-size: 14px; color: var(--ok); font-weight: 800; min-height: 18px; }
      .actions { display: flex; align-items: center; gap: 14px; margin-top: 18px; flex-wrap: wrap; }
      button, .button {
        border: none; border-radius: 14px; padding: 13px 19px; font-size: 15px; font-weight: 900;
        background: var(--accent); color: white; cursor: pointer; text-decoration: none;
        display: inline-flex; align-items: center; gap: 8px; transition: .18s ease;
      }
      button:hover, .button:hover { transform: translateY(-1px); box-shadow: 0 14px 24px rgba(17,24,39,.18); }
      button.secondary, .button.secondary { background: #475467; }
      button.outline, .button.outline { background: white; color: #111827; border: 1px solid var(--line); }
      .hint { color: var(--muted); font-size: 13px; }
      .templates {
        display: grid; grid-template-columns: repeat(2, 1fr); gap: 12px; margin: 0 0 18px;
      }
      .template-card {
        background: white; border: 1px solid var(--line); border-radius: 18px; padding: 16px;
        display: flex; justify-content: space-between; gap: 12px; align-items: center;
      }
      .template-card strong { display:block; font-size:14px; }
      .template-card span { display:block; color:var(--muted); font-size:12px; margin-top:4px; }
      .features { display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; padding: 18px 34px 30px; }
      .feature { background: white; border: 1px solid var(--evalia-border); border-radius: 18px; padding: 16px; color: var(--muted); font-size: 13px; box-shadow: 0 8px 24px rgba(14,165,233,.06); }
      .feature strong { display: block; color: var(--ink); font-size: 14px; margin-bottom: 5px; }
      .loader { display: none; color: var(--muted); font-size: 14px; margin-top: 16px; }
      .progress-container {
        width: 100%; height: 14px; background: #e5e7eb; border-radius: 999px;
        overflow: hidden; margin-top: 10px;
      }
      .progress-bar {
        width: 0%; height: 100%; border-radius: 999px;
        background: linear-gradient(90deg,var(--alt-blue),var(--alt-green),var(--alt-orange),var(--alt-purple));
        transition: width .5s ease;
      }
      .progress-status { margin-top: 9px; font-size: 13px; color: var(--muted); }
      .preview-box {
        display:none; background:white; border:1px solid var(--line); border-radius:18px;
        padding:16px; margin-top:16px; color:#344054; font-size:14px; line-height:1.55;
      }
      .preview-title { font-weight:900; color:#111827; margin-bottom:6px; }
      .preview-ok { color:#067647; font-weight:800; }
      .preview-warn { color:#b54708; font-weight:800; }
      .metric-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin: 24px 0; }
      .metric { border: 1px solid var(--evalia-border); border-radius: 18px; padding: 16px; background: linear-gradient(180deg,#ffffff 0%,#f3f9ff 100%); }
      .metric-value { font-size: 24px; font-weight: 900; }
      .metric-label { color: var(--muted); font-size: 13px; margin-top: 4px; }
      .error {
        color: #b42318; background: #fff4f2; border: 1px solid #fecdca;
        border-radius: 18px; padding: 18px;
      }
      .footer-altiora { text-align: center; margin-top: 26px; color: #667085; font-size: 13px; padding-bottom: 18px; }
      .footer-sub { margin-top: 4px; font-size: 12px; letter-spacing:.02em; }
      .hero, .result-card { position: relative; }
      .hero:before, .result-card:before {
        content:""; position:absolute; left:0; top:0; right:0; height:4px;
        background: linear-gradient(90deg,var(--alt-blue),var(--alt-green),var(--alt-orange),var(--alt-purple));
      }
      code { background: #eef2f7; padding: 2px 6px; border-radius: 6px; }
      @media (max-width: 760px) {
        h1 { font-size: 32px; }
        .features, .metric-grid, .templates { grid-template-columns: 1fr; }
        .topbar { align-items: flex-start; gap: 14px; flex-direction: column; }
      }
    .dropzone{border:2px dashed #93c5fd;background:#f8fbff;border-radius:16px;padding:14px;margin:8px 0 12px 0}.dropzone input{display:block;width:100%;margin-bottom:6px}.dropzone small{color:#475569}.ocr-warning{background:#fff7ed;border:1px solid #fed7aa;color:#7c2d12;padding:12px;border-radius:14px;margin:12px 0}</style>
    """


def shell_topbar(subtitle="Evalia by Altiora · Inteligencia Evaluativa Automatizada", badge="CRB Engine · v3.5 baseline validación"):
    return f"""
    <div class="topbar">
      <div class="brand">
        <div class="logo">E</div>
        <div>
          <div class="brand-title">Evalia</div>
          <div class="brand-subtitle">{escape(subtitle)}</div>
        </div>
      </div>
      <div class="badge">{escape(badge)}</div>
    </div>
    """


def footer_altiora():
    return """
    <footer class="footer-altiora">
      <div><strong>Evalia by Altiora</strong></div>
      <div class="footer-sub">Inteligencia que eleva posibilidades</div>
    </footer>
    """


def save_template_workbook(kind: str):
    wb = Workbook()
    ws = wb.active

    if kind == "rubric":
        ws.title = "Modelo_Rubrica"
        headers = ["Pregunta", "Tipo de pregunta", "Puntaje", "Ideas esperadas", "Enunciado"]
        ws.append(headers)
        rows = [
            ["P1", "Pregunta abierta", 5, "gravedad, órbita, Sol, planetas", "Explique por qué los planetas orbitan alrededor del Sol."],
            ["P2", "Verdadero/Falso", 1, "Verdadero", "La Luna es el satélite natural de la Tierra."],
            ["P3", "Completar", 2, "Vía Láctea", "Complete: El sistema solar pertenece a la ________."],
            ["P4", "Enumerar", 3, "Mercurio, Venus, Tierra, Marte", "Mencione al menos dos planetas rocosos del sistema solar."],
            ["P5", "Relacionar", 4, "Júpiter:gigante gaseoso; Marte:planeta rojo; Saturno:anillos; Neptuno:azul", "Relacione planeta y característica."],
            ["P6", "Pregunta abierta", 5, "estrella, energía, fusión nuclear, luz, calor", "Explique qué es una estrella y cómo produce energía."]
        ]
        for row in rows:
            ws.append(row)

        info = wb.create_sheet("Instrucciones")
        info.append(["Evalia", "Cómo completar el Modelo de Rúbrica"])
        info.append(["Pregunta", "Use códigos simples: P1, P2, P3..."])
        info.append(["Tipo de pregunta", "Use: Pregunta abierta, Verdadero/Falso, Completar, Enumerar o Relacionar."])
        info.append(["Puntaje", "Puntaje máximo de la pregunta."])
        info.append(["Ideas esperadas", "Escriba las ideas clave o la respuesta esperada. Puede usar comas, punto y coma o saltos de línea."])
        info.append(["Enunciado", "Pregunta o instrucción que vio el estudiante."])
        info.append(["Relacionar", "Para relaciones use: Concepto:Respuesta; Concepto:Respuesta."])
        info.append(["Nota", "No necesita escribir sinónimos técnicos. Evalia intenta reconocer paráfrasis, relaciones y contradicciones."])
        filename = "modelo_rubrica_evalia.xlsx"

    else:
        ws.title = "Modelo_Respuestas"
        headers = ["student_id", "nombre", "P1", "P2", "P3", "P4", "P5", "P6"]
        ws.append(headers)
        rows = [
            ["A01", "Ana Pérez", "Los planetas giran alrededor del Sol por la gravedad y el movimiento orbital.", "Verdadero", "Vía Láctea", "Tierra, Marte", "Júpiter gigante gaseoso; Marte planeta rojo", "Una estrella produce luz y calor mediante fusión nuclear."],
            ["A02", "Luis Soto", "La atracción del Sol mantiene a los cuerpos celestes en órbita.", "V", "nuestra galaxia", "Mercurio, Venus", "Saturno anillos; Neptuno azul", "Las estrellas emiten energía por reacciones nucleares."],
            ["A03", "Camila Díaz", "Porque el Sol los atrae.", "Falso", "Andrómeda", "Marte", "Júpiter azul", "Una estrella es un planeta caliente."]
        ]
        for row in rows:
            ws.append(row)

        info = wb.create_sheet("Instrucciones")
        info.append(["Evalia", "Cómo completar el Modelo de Respuestas"])
        info.append(["student_id", "Identificador del estudiante."])
        info.append(["nombre", "Nombre del estudiante."])
        info.append(["P1, P2, P3...", "Respuesta del estudiante para cada pregunta. Debe coincidir con los códigos de la rúbrica."])
        info.append(["Nota", "Se recomienda usar P1, P2, P3... para evitar errores."])
        filename = "modelo_respuestas_evalia.xlsx"

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    soft_fill = PatternFill("solid", fgColor="F9FAFB")

    for sheet in wb.worksheets:
        if sheet.max_row >= 1:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row % 2 == 0:
                    cell.fill = soft_fill

        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 34

    path = OUTPUT_DIR / filename
    wb.save(path)
    return path

@app.get("/download-template/rubric")
def download_rubric_template():
    path = save_template_workbook("rubric")
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=path.name
    )


@app.get("/download-template/responses")
def download_response_template():
    path = save_template_workbook("responses")
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=path.name
    )


@app.post("/preview")
async def preview_upload(
    file: UploadFile = File(...),
    rubric_selector: str = Form(""),
    rubric_file: Optional[UploadFile] = File(None)
):
    try:
        if rubric_file is not None and rubric_file.filename:
            selected_rubric = await load_uploaded_rubric(rubric_file)
        else:
            raise ValueError("Debes subir el Modelo de Rúbrica para previsualizar.")

        temp_input_path = OUTPUT_DIR / f"preview_{file.filename}"
        with open(temp_input_path, "wb") as f:
            f.write(await file.read())

        df = pd.read_excel(temp_input_path)
        questions = selected_rubric.get("questions", [])
        rubric_issues = validate_rubric_integrity(selected_rubric)
        df_issues = validate_dataframe_integrity(df)
        missing = validate_columns_flexible(df, selected_rubric)
        format_issues = rubric_issues + df_issues

        detected = []
        for q in questions:
            pid = q.get("id", "")
            col = find_item_column(df, pid)
            detected.append({
                "pregunta": pid,
                "tipo": display_item_type(q.get("item_type", "")),
                "columna_detectada": col or "NO DETECTADA"
            })

        type_counts = {}
        for q in questions:
            t = display_item_type(q.get("item_type", "sin_tipo"))
            type_counts[t] = type_counts.get(t, 0) + 1

        return JSONResponse({
            "ok": len(missing) == 0 and len(format_issues) == 0,
            "rubric": selected_rubric.get("_rubric_display_name", selected_rubric.get("name", "Rúbrica")),
            "students": int(len(df)),
            "questions": int(len(questions)),
            "columns": [str(c) for c in df.columns],
            "types": type_counts,
            "detected": detected,
            "missing": missing + format_issues
        })

    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)




# ============================================================
# EVALIA OCR-MVP v1 · CAPA DE INTELIGENCIA EVALUATIVA EXPLICABLE
# ============================================================
# Esta capa NO reemplaza el motor semántico v3.5: lo reutiliza.
# Agrega flujo por estudiante: imágenes -> OCR -> segmentación -> edición docente -> evaluación.

IMAGE_DIR = OUTPUT_DIR / "ocr_images"
IMAGE_DIR.mkdir(exist_ok=True)

OCR_STATUS = {
    "engine": "manual_fallback",
    "available": False,
    "message": "OCR automático no disponible; se habilita transcripción/edición manual."
}

try:
    from PIL import Image, ImageOps, ImageFilter
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False


def safe_session_id(*parts):
    raw = "||".join(str(x) for x in parts) + "||" + str(time.time())
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:12]


def normalize_question_id_for_regex(qid):
    n = item_number(qid)
    if n:
        return [re.escape(str(qid)), rf"p\s*{n}", rf"pregunta\s*{n}", rf"item\s*{n}", rf"{n}\s*[\)\.-]"]
    return [re.escape(str(qid))]


def preprocess_image_for_ocr(path):
    """Preprocesamiento liviano para fotos WhatsApp: orientación, escala, gris y contraste."""
    if not PIL_AVAILABLE:
        return path
    try:
        img = Image.open(path)
        img = ImageOps.exif_transpose(img)
        max_side = max(img.size)
        if max_side < 1800:
            scale = 1800 / max_side
            img = img.resize((int(img.width * scale), int(img.height * scale)))
        img = ImageOps.grayscale(img)
        img = ImageOps.autocontrast(img)
        img = img.filter(ImageFilter.SHARPEN)
        out = Path(str(path) + "_pre.png")
        img.save(out)
        return out
    except Exception as e:
        log_event("ocr_preprocess_failed", file=path.name, error=str(e))
        return path


def run_ocr_on_image(path):
    """OCR híbrido seguro v1.1: intenta pytesseract y conserva líneas/bloques.
    Si el binario de Tesseract no está disponible en Render, activa fallback manual transparente.
    """
    started = time.time()
    if not PIL_AVAILABLE:
        return {
            "text": "",
            "confidence": 0.0,
            "engine": "manual_fallback",
            "seconds": 0.0,
            "message": "PIL no está disponible; transcripción manual requerida."
        }
    try:
        import pytesseract
        img_path = preprocess_image_for_ocr(path)
        img = Image.open(img_path)
        config = os.getenv("EVALIA_TESSERACT_CONFIG", "--oem 3 --psm 6")
        lang = os.getenv("EVALIA_OCR_LANG", "spa+eng")
        data = pytesseract.image_to_data(img, lang=lang, config=config, output_type=pytesseract.Output.DICT)

        rows = []
        confs = []
        n = len(data.get("text", []))
        for i in range(n):
            txt = str(data.get("text", [""])[i]).strip()
            if not txt:
                continue
            try:
                c = float(data.get("conf", [0])[i])
                if c >= 0:
                    confs.append(c)
            except Exception:
                c = 0
            rows.append({
                "block": data.get("block_num", [0])[i],
                "par": data.get("par_num", [0])[i],
                "line": data.get("line_num", [0])[i],
                "word": data.get("word_num", [0])[i],
                "left": data.get("left", [0])[i],
                "top": data.get("top", [0])[i],
                "text": txt,
                "conf": c,
            })

        # Reconstrucción respetando líneas. Esto es clave para segmentar respuestas.
        grouped = {}
        for r in rows:
            key = (r["block"], r["par"], r["line"])
            grouped.setdefault(key, []).append(r)
        lines = []
        for key in sorted(grouped.keys()):
            words = sorted(grouped[key], key=lambda x: (x["left"], x["word"]))
            line = " ".join(w["text"] for w in words).strip()
            if line:
                lines.append(line)
        text = "\n".join(lines).strip()
        avg_conf = round((sum(confs) / len(confs)) / 100, 2) if confs else 0.0

        if not text:
            OCR_STATUS.update({"engine": "pytesseract_empty", "available": False, "message": "OCR ejecutado sin texto útil."})
            return {
                "text": "",
                "confidence": 0.0,
                "engine": "pytesseract_empty",
                "seconds": round(time.time() - started, 3),
                "message": "OCR ejecutado, pero no se detectó texto útil. Usa el campo de transcripción manual."
            }

        OCR_STATUS.update({"engine": "pytesseract", "available": True, "message": "OCR automático activo con pytesseract."})
        return {
            "text": text,
            "confidence": avg_conf,
            "engine": "pytesseract_lines",
            "seconds": round(time.time() - started, 3),
            "message": "OCR automático aplicado conservando líneas."
        }
    except Exception as e:
        msg = str(e)
        hint = ""
        if "tesseract" in msg.lower():
            hint = " En Render esto suele indicar que falta el binario del sistema Tesseract, no solo la librería pytesseract."
        log_event("ocr_unavailable_or_failed", file=path.name, error=msg)
        return {
            "text": "",
            "confidence": 0.0,
            "engine": "manual_fallback",
            "seconds": round(time.time() - started, 3),
            "message": f"OCR no disponible o falló: {msg}.{hint} Puedes pegar/transcribir el texto manualmente."
        }

def compact_for_matching(txt):
    return re.sub(r"\s+", " ", normalize_text(txt or "")).strip()


def strip_prompt_from_segment(segment, prompt):
    """Quita el enunciado detectado dentro del OCR para dejar preferentemente la respuesta."""
    seg = str(segment or "").strip()
    pr = str(prompt or "").strip()
    if not seg or not pr:
        return seg
    seg_norm = compact_for_matching(seg)
    pr_norm = compact_for_matching(pr)
    if not pr_norm:
        return seg
    # Si el prompt aparece casi literal, corta después del prompt original aproximado por longitud.
    pos = seg_norm.find(pr_norm[:max(18, min(len(pr_norm), 55))])
    if pos >= 0:
        # Fallback simple: remueve las palabras del prompt desde el inicio si están cerca.
        prompt_words = pr_norm.split()
        seg_words = seg.split()
        if len(seg_words) > len(prompt_words):
            return " ".join(seg_words[min(len(prompt_words), len(seg_words)):]).strip()
    return seg


def segment_ocr_text_by_questions(raw_text, rubric):
    """Segmentador OCR v1.1.
    Usa tres capas: marcas numéricas, anclas por enunciado de rúbrica y fallback manual.
    """
    text = str(raw_text or "").replace("\r", "\n")
    questions = rubric.get("questions", [])
    segments = {str(q.get("id", "")): "" for q in questions}
    if not text.strip() or not questions:
        return segments, {"mode": "empty_or_manual", "confidence": 0.0, "notes": "No hay texto OCR suficiente para segmentar."}

    # Normaliza marcas frecuentes que Tesseract confunde.
    working = re.sub(r"(?m)^\s*(\d+)\s+", r"\1. ", text)
    working = re.sub(r"(?i)preg\.?\s*(\d+)", r"Pregunta \1", working)

    markers = []
    for q in questions:
        qid = str(q.get("id", ""))
        for pat in normalize_question_id_for_regex(qid):
            for m in re.finditer(rf"(?im)(^|\n|\s)({pat})(\s|:)", working):
                markers.append((m.start(2), qid, m.end(2), "marker"))
                break
            if any(x[1] == qid for x in markers):
                break

    # Segunda capa: buscar inicios por fragmentos del prompt.
    low_work = compact_for_matching(working)
    for q in questions:
        qid = str(q.get("id", ""))
        if any(x[1] == qid for x in markers):
            continue
        prompt = compact_for_matching(q.get("prompt", ""))
        if len(prompt) < 18:
            continue
        candidates = [prompt[:60], prompt[:40], " ".join(prompt.split()[:6])]
        for cand in candidates:
            if len(cand) < 18:
                continue
            idx = low_work.find(cand)
            if idx >= 0:
                # Índice aproximado sobre texto original; suficiente para ordenar cortes.
                approx = int(idx / max(len(low_work), 1) * len(working))
                markers.append((approx, qid, approx, "prompt_anchor"))
                break

    # Deduplicar por pregunta, conservando la primera posición.
    dedup = {}
    for pos, qid, endpos, mode in sorted(markers, key=lambda x: x[0]):
        dedup.setdefault(qid, (pos, qid, endpos, mode))
    markers = sorted(dedup.values(), key=lambda x: x[0])

    if len(markers) >= 2:
        q_by_id = {str(q.get("id", "")): q for q in questions}
        for i, (start, qid, end_marker, mode) in enumerate(markers):
            end = markers[i + 1][0] if i + 1 < len(markers) else len(working)
            seg = working[end_marker:end]
            seg = re.sub(r"^\s*[:\).-]*\s*", "", seg).strip()
            seg = strip_prompt_from_segment(seg, q_by_id.get(qid, {}).get("prompt", ""))
            # Limpieza de restos de corrección docente frecuentes.
            seg = re.sub(r"(?i)\b(bien|buena justificacion|buena justificación|falto mas detalle|faltó más detalle)\b", "", seg).strip()
            segments[qid] = seg
        filled = len([v for v in segments.values() if v.strip()])
        conf = min(0.92, 0.40 + 0.08 * filled + (0.10 if any(m[3] == "prompt_anchor" for m in markers) else 0))
        return segments, {"mode": "markers_plus_prompt_anchors", "confidence": round(conf, 2), "notes": "Segmentación por marcas visibles y/o enunciados de la rúbrica."}

    # Fallback: si el docente pegó texto ordenado con separadores tipo P1:, P2:, etc.
    manual_markers = re.split(r"(?im)(?:^|\n)\s*(?:p|pregunta)?\s*(\d{1,2})\s*[:\).-]\s*", working)
    if len(manual_markers) > 2:
        for i in range(1, len(manual_markers), 2):
            n = manual_markers[i]
            ans = manual_markers[i+1] if i+1 < len(manual_markers) else ""
            for q in questions:
                if item_number(q.get("id")) == n:
                    segments[str(q.get("id"))] = ans.strip()
        return segments, {"mode": "manual_numbered_text", "confidence": 0.70, "notes": "Segmentación desde texto numerado pegado/editado manualmente."}

    first = str(questions[0].get("id", ""))
    segments[first] = working.strip()
    return segments, {"mode": "manual_review_required", "confidence": 0.25, "notes": "No se detectaron marcas suficientes; revisar y distribuir respuestas manualmente."}

def ocr_confidence_state(ocr_conf, seg_conf):
    combined = round((0.65 * float(ocr_conf or 0)) + (0.35 * float(seg_conf or 0)), 2)
    if combined >= 0.78:
        return combined, "verde", "OCR/segmentación confiable"
    if combined >= 0.48:
        return combined, "amarillo", "OCR/segmentación aceptable con cautela"
    return combined, "rojo", "Revisión manual recomendada"


def cognitive_level_from_score(score, max_score, confidence, diagnosis):
    pct = (float(score) / float(max_score or 1)) if max_score else 0
    coverage = float(diagnosis.get("conceptual_coverage", 0) or 0)
    relations = bool(diagnosis.get("conceptual_relations", ""))
    profile = diagnosis.get("answer_profile", "")
    if pct <= 0.05:
        return "E0 · incorrecto/no evaluable"
    if pct < 0.40 or coverage < 0.30:
        return "E1 · fragmentario"
    if pct < 0.70:
        return "E2 · funcional"
    if pct < 0.90 or not relations:
        return "E3 · elaborado"
    if relations and profile == "developed" and confidence >= 0.75:
        return "E4 · profundo/inferencial"
    return "E3 · elaborado"


def hidden_input(name, value):
    return f'<input type="hidden" name="{escape(name)}" value="{escape(str(value), quote=True)}">'


@app.get("/ocr", response_class=HTMLResponse)
def ocr_home():
    return HTMLResponse(f"""
    <!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Evalia OCR v1.1</title>{base_css()}</head>
    <body><div class="page"><main class="shell">
      {shell_topbar("Evalia OCR v1.1", "Inteligencia evaluativa explicable")}
      <section class="hero"><div class="hero-inner">
        <h1>Evaluación desde imágenes manuscritas</h1>
        <p class="lead">Sube una rúbrica, identifica al estudiante y carga fotos de sus hojas. Evalia hará OCR cuando esté disponible, segmentará respuestas y dejará todo editable antes de evaluar.</p>
        <form action="/ocr/process" enctype="multipart/form-data" method="post">
          <div class="panel">
            <label class="field-label">Curso</label><input name="course" placeholder="Ej.: Psicolingüística" style="width:100%;padding:12px;border-radius:14px;border:1px solid #d1d5db;">
            <label class="field-label">Certamen / evaluación</label><input name="exam_name" placeholder="Ej.: Certamen 1" style="width:100%;padding:12px;border-radius:14px;border:1px solid #d1d5db;">
            <label class="field-label">Fecha</label><input name="exam_date" placeholder="2026-05-08" style="width:100%;padding:12px;border-radius:14px;border:1px solid #d1d5db;">
            <label class="field-label">ID estudiante</label><input name="student_id" required placeholder="Ej.: A01" style="width:100%;padding:12px;border-radius:14px;border:1px solid #d1d5db;">
            <label class="field-label">Nombre estudiante</label><input name="student_name" required placeholder="Nombre completo" style="width:100%;padding:12px;border-radius:14px;border:1px solid #d1d5db;">
            <label class="field-label">Rúbrica Excel/JSON</label><div class="dropzone"><input name="rubric_file" type="file" accept=".xlsx,.xls,.json" required><small>Arrastra o selecciona la rúbrica Excel/JSON</small></div>
            <label class="field-label">Imágenes de hojas del estudiante</label><div class="dropzone"><input name="image_files" type="file" accept="image/*" multiple required><small>Arrastra o selecciona una o más fotos/escaneos del estudiante</small></div>
            <div class="actions"><button type="submit">Procesar OCR y revisar</button><a class="button secondary" href="/">Volver a Excel</a></div>
          </div>
        </form>
      </div></section>{footer_altiora()}
    </main></div></body></html>
    """)


@app.post("/ocr/process", response_class=HTMLResponse)
async def ocr_process(
    course: str = Form(""),
    exam_name: str = Form(""),
    exam_date: str = Form(""),
    student_id: str = Form(""),
    student_name: str = Form(""),
    rubric_file: UploadFile = File(...),
    image_files: List[UploadFile] = File(...)
):
    try:
        rubric = await load_uploaded_rubric(rubric_file)
        issues = validate_rubric_integrity(rubric)
        if issues:
            return safe_error_page("Rúbrica con problemas", "La rúbrica debe corregirse antes de evaluar.", "; ".join(issues))

        session_id = safe_session_id(student_id, student_name, exam_name)
        saved_rubric = OUTPUT_DIR / f"ocr_session_{session_id}_rubric.json"
        saved_rubric.write_text(json.dumps(rubric, ensure_ascii=False), encoding="utf-8")

        ocr_blocks = []
        all_text_parts = []
        for idx, img in enumerate(image_files, start=1):
            suffix = Path(img.filename or f"hoja_{idx}.png").suffix or ".png"
            safe_name = f"{session_id}_hoja_{idx}{suffix}"
            out_path = IMAGE_DIR / safe_name
            with open(out_path, "wb") as f:
                f.write(await img.read())
            result = run_ocr_on_image(out_path)
            ocr_blocks.append({"file": img.filename, "stored_file": safe_name, **result})
            if result.get("text"):
                all_text_parts.append(f"\n\n[HOJA {idx}]\n" + result.get("text", ""))

        raw_text = "\n".join(all_text_parts).strip()
        segments, seg_info = segment_ocr_text_by_questions(raw_text, rubric)
        avg_ocr_conf = round(sum(float(b.get("confidence", 0) or 0) for b in ocr_blocks) / max(len(ocr_blocks), 1), 2)
        combined_conf, color_state, state_label = ocr_confidence_state(avg_ocr_conf, seg_info.get("confidence", 0))

        meta = {"course": course, "exam_name": exam_name, "exam_date": exam_date, "student_id": student_id, "student_name": student_name}
        saved_meta = OUTPUT_DIR / f"ocr_session_{session_id}_meta.json"
        saved_meta.write_text(json.dumps({"meta": meta, "ocr_blocks": ocr_blocks, "raw_text": raw_text, "segmentation": seg_info}, ensure_ascii=False), encoding="utf-8")

        q_html = []
        for q in rubric.get("questions", []):
            qid = str(q.get("id", ""))
            prompt = q.get("prompt", "")
            q_html.append(f"""
            <div class="template-card" style="display:block;">
              <strong>{escape(qid)} · {escape(display_item_type(q.get('item_type','')))} · {q.get('max_score', '')} pts</strong>
              <span>{escape(prompt)}</span>
              <textarea name="answer__{escape(qid)}" rows="5" style="width:100%;margin-top:10px;padding:12px;border-radius:14px;border:1px solid #d1d5db;">{escape(segments.get(qid, ''))}</textarea>
            </div>
            """)

        ocr_detail = "<br>".join([escape(f"{b.get('file')}: {b.get('engine')} · conf={b.get('confidence')} · {b.get('message')}") for b in ocr_blocks])
        color_badge = {"verde":"#dcfce7", "amarillo":"#fef9c3", "rojo":"#fee2e2"}.get(color_state, "#f3f4f6")

        return HTMLResponse(f"""
        <!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Revisión OCR · Evalia</title>{base_css()}</head>
        <body><div class="page"><main class="shell">
          {shell_topbar("Revisión docente OCR", "Evalia OCR v1.1")}
          <section class="result-card">
            <h1>Revisa y corrige antes de evaluar</h1>
            <p class="lead"><strong>{escape(student_name)}</strong> · {escape(student_id)} · {escape(course)} · {escape(exam_name)}</p>
            <div class="metric-grid">
              <div class="metric"><div class="metric-value">{combined_conf}</div><div class="metric-label">confianza OCR global</div></div>
              <div class="metric"><div class="metric-value">{seg_info.get('confidence',0)}</div><div class="metric-label">confianza segmentación</div></div>
              <div class="metric"><div class="metric-value">{len(rubric.get('questions', []))}</div><div class="metric-label">preguntas</div></div>
            </div>
            <div style="background:{color_badge};padding:12px;border-radius:14px;margin:12px 0;"><strong>{escape(state_label)}</strong><br><small>{escape(seg_info.get('notes',''))}</small><br><small>{ocr_detail}</small></div>
            <form action="/ocr/evaluate" method="post">
              {hidden_input('session_id', session_id)}
              {''.join(q_html)}
              <label class="field-label">Texto OCR bruto editable / respaldo</label>
              <textarea name="raw_text_edited" rows="8" style="width:100%;padding:12px;border-radius:14px;border:1px solid #d1d5db;">{escape(raw_text)}</textarea>
              <div class="actions"><button type="submit">Evaluar respuestas editadas</button><a class="button secondary" href="/ocr">Cargar otro estudiante</a></div>
            </form>
          </section>{footer_altiora()}
        </main></div></body></html>
        """)
    except Exception as e:
        logger.error("ocr_process_failed | %s\n%s", e, traceback.format_exc())
        return safe_error_page("Error OCR controlado", "No se pudo completar el procesamiento OCR.", str(e))


@app.post("/ocr/evaluate", response_class=HTMLResponse)
async def ocr_evaluate(request: Request):
    try:
        form = await request.form()
        session_id = str(form.get("session_id", ""))
        rubric_path = OUTPUT_DIR / f"ocr_session_{session_id}_rubric.json"
        meta_path = OUTPUT_DIR / f"ocr_session_{session_id}_meta.json"
        if not rubric_path.exists() or not meta_path.exists():
            return safe_error_page("Sesión no encontrada", "No se encontró la sesión OCR. Vuelve a cargar las imágenes.")
        rubric = json.loads(rubric_path.read_text(encoding="utf-8"))
        meta_payload = json.loads(meta_path.read_text(encoding="utf-8"))
        meta = meta_payload.get("meta", {})

        score_rows, feedback_rows, trace_rows, semantic_rows = [], [], [], []
        total = 0.0
        total_score = float(rubric.get("total_score", 0)) or sum(float(q.get("max_score", 0)) for q in rubric.get("questions", [])) or 1
        accepted = caution = review = 0

        row_score = {"student_id": meta.get("student_id", ""), "nombre": meta.get("student_name", "")}
        row_conf = {"student_id": meta.get("student_id", ""), "nombre": meta.get("student_name", "")}

        for q in rubric.get("questions", []):
            qid = str(q.get("id", ""))
            answer = str(form.get(f"answer__{qid}", ""))
            if not answer.strip() and fallback_segments.get(qid, "").strip():
                answer = fallback_segments.get(qid, "")
            score, conf, fb, status = score_answer(answer, q)
            diagnosis = semantic_diagnosis(answer, q, score=score, confidence=conf, status=status)
            cog_level = cognitive_level_from_score(score, q.get("max_score", 1), conf, diagnosis)
            total += float(score)
            row_score[qid] = score
            row_conf[qid] = conf
            if status == "aceptado": accepted += 1
            elif status == "aceptado_con_cautela": caution += 1
            else: review += 1
            feedback_rows.append({
                "student_id": meta.get("student_id", ""), "nombre": meta.get("student_name", ""), "pregunta": qid,
                "respuesta_editada": answer, "puntaje": score, "puntaje_maximo": q.get("max_score", ""),
                "confianza_evaluativa": conf, "estado": status, "nivel_cognitivo": cog_level, "feedback": fb
            })
            trace_rows.append(build_traceability_row(meta.get("student_id", ""), meta.get("student_name", ""), qid, answer, score, conf, status, diagnosis, fb))
            sem = {"student_id": meta.get("student_id", ""), "nombre": meta.get("student_name", ""), "pregunta": qid, **diagnosis}
            semantic_rows.append(sem)

        pct = round((total / total_score) * 100, 2)
        row_score.update({"total": round(total, 2), "porcentaje": pct, "nivel_desempeno": performance_level(pct)})
        row_conf.update({"confianza_promedio": round(sum([float(r.get('confianza_evaluativa',0)) for r in feedback_rows]) / max(len(feedback_rows),1), 2)})
        score_rows.append(row_score)

        output_name = f"evalia_ocr_{session_id}_{normalize_column_name(meta.get('student_name','estudiante'))}.xlsx"
        output_path = OUTPUT_DIR / output_name
        n = max(len(rubric.get("questions", [])), 1)
        summary = [{
            "version": APP_VERSION,
            "curso": meta.get("course", ""), "evaluacion": meta.get("exam_name", ""), "fecha": meta.get("exam_date", ""),
            "student_id": meta.get("student_id", ""), "nombre": meta.get("student_name", ""),
            "puntaje_total": round(total, 2), "puntaje_maximo": total_score, "porcentaje": pct,
            "nivel_desempeno": performance_level(pct),
            "aceptado_pct": round(accepted / n * 100, 1), "cautela_pct": round(caution / n * 100, 1), "revision_pct": round(review / n * 100, 1),
            "ocr_engine": OCR_STATUS.get("engine"), "criterio_mvp": "copiloto evaluativo: OCR editable + microjuicios explicables + revisión docente"
        }]

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            pd.DataFrame(summary).to_excel(writer, sheet_name="RESUMEN_OCR", index=False)
            pd.DataFrame(score_rows).to_excel(writer, sheet_name="Puntajes", index=False)
            pd.DataFrame([row_conf]).to_excel(writer, sheet_name="Confianza", index=False)
            pd.DataFrame(feedback_rows).to_excel(writer, sheet_name="Feedback", index=False)
            pd.DataFrame(trace_rows).to_excel(writer, sheet_name="TRAZABILIDAD_EVALIA", index=False)
            pd.DataFrame(semantic_rows).to_excel(writer, sheet_name="MICROJUICIOS", index=False)
            pd.DataFrame(meta_payload.get("ocr_blocks", [])).to_excel(writer, sheet_name="OCR_HOJAS", index=False)
            pd.DataFrame([{"raw_text_edited": str(form.get("raw_text_edited", "")), "raw_text_original": meta_payload.get("raw_text", "")}]).to_excel(writer, sheet_name="OCR_TEXTO", index=False)
            format_workbook(writer)

        return HTMLResponse(f"""
        <!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Resultado OCR · Evalia</title>{base_css()}</head>
        <body><div class="page"><main class="shell">
          {shell_topbar("Reporte OCR generado", "Evalia OCR v1.1")}
          <section class="result-card">
            <h1>Evaluación completada</h1>
            <p class="lead"><strong>{escape(meta.get('student_name',''))}</strong> obtuvo {round(total,2)} / {total_score} puntos ({pct}%).</p>
            <div class="metric-grid">
              <div class="metric"><div class="metric-value">{performance_level(pct)}</div><div class="metric-label">desempeño global</div></div>
              <div class="metric"><div class="metric-value">{round(accepted/n*100,1)}%</div><div class="metric-label">aceptado</div></div>
              <div class="metric"><div class="metric-value">{round(caution/n*100,1)}%</div><div class="metric-label">cautela</div></div>
              <div class="metric"><div class="metric-value">{round(review/n*100,1)}%</div><div class="metric-label">revisión</div></div>
            </div>
            <div class="actions"><a class="button" href="/download/{output_name}">Descargar Excel OCR</a><a class="button secondary" href="/ocr">Evaluar otro estudiante</a></div>
          </section>{footer_altiora()}
        </main></div></body></html>
        """)
    except Exception as e:
        logger.error("ocr_evaluate_failed | %s\n%s", e, traceback.format_exc())
        return safe_error_page("Error de evaluación OCR", "No se pudo evaluar la sesión OCR.", str(e))


@app.get("/", response_class=HTMLResponse)
def home():
    html = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
      <meta charset="UTF-8">
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
      <title>Evalia by Altiora · Inteligencia Evaluativa Automatizada</title>
      {base_css()}
    </head>
    <body>
      <div class="page">
        <main class="shell">
          {shell_topbar()}

          <div class="template-card" style="margin-bottom:18px;">
            <div><strong>Nuevo flujo OCR-MVP v1</strong><span>Evaluar hojas manuscritas por estudiante con OCR editable y microjuicios explicables.</span></div>
            <a class="button" href="/ocr">Abrir OCR-MVP</a>
          </div>

          <section class="hero">
            <div class="hero-inner">
              <h1>Evalúa respuestas. Detecta patrones. Mejora evaluaciones.</h1>
              <p class="lead">
                Evalia trabaja de forma simple: descarga el Modelo de Rúbrica y el Modelo de Respuestas,
                complétalos y sube ambos archivos para generar un reporte docente explicable.
              </p>

              <div class="templates">
                <div class="template-card">
                  <div>
                    <strong>Modelo de Rúbrica</strong>
                    <span>Formato simple para docentes: pregunta, tipo de pregunta, puntaje, ideas esperadas y enunciado.</span>
                  </div>
                  <a class="button outline" href="/download-template/rubric">Descargar</a>
                </div>
                <div class="template-card">
                  <div>
                    <strong>Modelo de Respuestas</strong>
                    <span>Formato simple para cargar estudiantes y respuestas.</span>
                  </div>
                  <a class="button outline" href="/download-template/responses">Descargar</a>
                </div>
              </div>

              <form action="/upload" enctype="multipart/form-data" method="post" id="uploadForm">
                <div class="panel">

                  <label class="field-label">1. Sube tu rúbrica</label>
                  <div class="dropzone">
                    <input name="rubric_file" id="rubricFileInput" type="file" accept=".xlsx,.xls,.json" required>
                    <div class="drop-icon">🧩</div>
                    <div class="drop-title">Sube tu rúbrica</div>
                    <div class="drop-subtitle">Usa el formato simple descargable</div>
                    <div class="file-name" id="rubricFileName"></div>
                  </div>

                  <label class="field-label">2. Sube las respuestas</label>
                  <div class="dropzone">
                    <input name="file" id="fileInput" type="file" accept=".xlsx,.xls" required>
                    <div class="drop-icon">📄</div>
                    <div class="drop-title">Sube las respuestas</div>
                    <div class="drop-subtitle">Usa el formato simple descargable</div>
                    <div class="file-name" id="fileName"></div>
                  </div>

                  <div class="actions">
                    <button type="button" class="secondary" id="previewBtn">Previsualizar</button>
                    <button type="submit" id="submitBtn">Evaluar respuestas</button>
                    <span class="hint">Evalia analizará la rúbrica y las respuestas que subas.</span>
                  </div>

                  <div class="preview-box" id="previewBox"></div>

                  <div class="loader" id="loader">
                    <div class="progress-container"><div class="progress-bar" id="progressBar"></div></div>
                    <div class="progress-status" id="progressStatus">Preparando análisis...</div>
                  </div>
                </div>
              </form>
            </div>

            <div class="features">
              <div class="feature"><strong>Modelo de Rúbrica</strong>Define preguntas, puntajes y criterios.</div>
              <div class="feature"><strong>Modelo de Respuestas</strong>Organiza estudiantes y respuestas.</div>
              <div class="feature"><strong>Vista previa</strong>Confirma que el formato calza.</div>
              <div class="feature"><strong>Capa semántica</strong>Reconoce paráfrasis, relaciones y posibles contradicciones conceptuales.</div>
            </div>
          </section>

          {footer_altiora()}
        </main>
      </div>

      <script>
        function bindFile(inputId, labelId) {{
          const input = document.getElementById(inputId);
          const label = document.getElementById(labelId);
          if (input) {{
            input.addEventListener("change", () => {{
              if (input.files.length > 0) {{
                label.textContent = "Archivo seleccionado: " + input.files[0].name;
              }}
            }});
          }}
        }}

        bindFile("fileInput", "fileName");
        bindFile("rubricFileInput", "rubricFileName");

        const form = document.getElementById("uploadForm");
        const loader = document.getElementById("loader");
        const submitBtn = document.getElementById("submitBtn");
        const previewBtn = document.getElementById("previewBtn");
        const previewBox = document.getElementById("previewBox");

        const steps = [
          [18, "Leyendo Excel..."],
          [38, "Detectando preguntas y columnas..."],
          [62, "Evaluando respuestas..."],
          [82, "Generando insights..."],
          [100, "Construyendo reporte docente..."]
        ];
        let idx = 0;

        function simulateProgress() {{
          if (idx >= steps.length) return;
          const [pct, txt] = steps[idx];
          const bar = document.getElementById("progressBar");
          const status = document.getElementById("progressStatus");
          if (bar) bar.style.width = pct + "%";
          if (status) status.innerText = txt;
          idx++;
          setTimeout(simulateProgress, 700);
        }}

        if (previewBtn) {{
          previewBtn.addEventListener("click", async () => {{
            const fileInput = document.getElementById("fileInput");
            if (!fileInput.files.length) {{
              previewBox.style.display = "block";
              previewBox.innerHTML = '<span class="preview-warn">Primero sube el archivo de respuestas.</span>';
              return;
            }}

            const data = new FormData(form);
            previewBtn.disabled = true;
            previewBtn.textContent = "Previsualizando...";

            try {{
              const res = await fetch("/preview", {{ method: "POST", body: data }});
              const json = await res.json();

              previewBox.style.display = "block";

              if (!json.ok) {{
                const msg = json.error ? json.error : ("Faltan columnas: " + (json.missing || []).join(", "));
                previewBox.innerHTML = '<div class="preview-title">Vista previa con observaciones</div><span class="preview-warn">' + msg + '</span>';
              }} else {{
                const types = Object.entries(json.types || {{}}).map(([k,v]) => k + ": " + v).join(" · ");
                const detected = (json.detected || []).map(x => x.pregunta + " → " + x.columna_detectada).join(" · ");
                previewBox.innerHTML =
                  '<div class="preview-title">Vista previa detectada</div>' +
                  '<span class="preview-ok">Formato compatible.</span><br>' +
                  '<strong>Rúbrica:</strong> ' + json.rubric + '<br>' +
                  '<strong>Estudiantes:</strong> ' + json.students + '<br>' +
                  '<strong>Preguntas:</strong> ' + json.questions + '<br>' +
                  '<strong>Tipos:</strong> ' + types + '<br>' +
                  '<strong>Columnas detectadas:</strong> ' + detected;
              }}
            }} catch (err) {{
              previewBox.style.display = "block";
              previewBox.innerHTML = '<span class="preview-warn">No fue posible generar la vista previa.</span>';
            }}

            previewBtn.disabled = false;
            previewBtn.textContent = "Previsualizar";
          }});
        }}

        if (form) {{
          form.addEventListener("submit", () => {{
            if (loader) loader.style.display = "block";
            idx = 0;
            simulateProgress();
            if (submitBtn) {{
              submitBtn.disabled = true;
              submitBtn.textContent = "Procesando...";
            }}
          }});
        }}
      </script>
    </body>
    </html>
    """
    return HTMLResponse(html)


@app.post("/upload")
async def upload(
    file: UploadFile = File(...),
    rubric_selector: str = Form(""),
    rubric_file: Optional[UploadFile] = File(None)
):
    try:
        if rubric_file is not None and rubric_file.filename:
            selected_rubric = await load_uploaded_rubric(rubric_file)
        else:
            raise ValueError("Debes subir el Modelo de Rúbrica para previsualizar.")
    except Exception as e:
        return HTMLResponse(
            f"""
            <!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Error · Evalia</title>{base_css()}</head>
            <body><div class="page"><main class="shell">{shell_topbar("Error de carga", "v2.4 docente")}<div class="result-card">
              <div class="error"><strong>Error al cargar el Modelo de Rúbrica.</strong><br>{escape(str(e))}</div>
              <br><a class="button" href="/">Volver</a>
            </div>{footer_altiora()}</main></div></body></html>
            """
        )

    input_path = OUTPUT_DIR / file.filename
    with open(input_path, "wb") as f:
        f.write(await file.read())

    try:
        df = pd.read_excel(input_path)
    except Exception as e:
        return HTMLResponse(
            f"""
            <!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Error · Evalia</title>{base_css()}</head>
            <body><div class="page"><main class="shell">{shell_topbar("Error de lectura", "v2.4 docente")}<div class="result-card">
              <div class="error"><strong>No se pudo leer el Excel.</strong><br>{escape(str(e))}</div>
              <br><a class="button" href="/">Volver</a>
            </div>{footer_altiora()}</main></div></body></html>
            """
        )

    rubric_issues = validate_rubric_integrity(selected_rubric)
    df_issues = validate_dataframe_integrity(df)
    missing = validate_columns_flexible(df, selected_rubric) + rubric_issues + df_issues

    if missing:
        return HTMLResponse(
            f"""
            <!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Error de formato · Evalia</title>{base_css()}</head>
            <body><div class="page"><main class="shell">{shell_topbar("Error de formato", "v2.4 docente")}<div class="result-card">
              <h1>Error de formato</h1>
              <div class="error"><strong>Faltan columnas requeridas o equivalentes:</strong><br>{escape(", ".join(missing))}</div>
              <p class="lead">Evalia acepta equivalencias como <code>P1/Q1/pregunta1/item1</code>, pero no encontró columnas suficientes.</p>
              <a class="button" href="/">Volver</a>
            </div>{footer_altiora()}</main></div></body></html>
            """
        )

    started_at = time.time()
    log_event("upload_started", file=file.filename, students=len(df), questions=len(selected_rubric.get("questions", [])))

    # v3.4: precálculo batch de embeddings para reducir llamadas repetidas al modelo.
    embedding_batch_info = precompute_embedding_vectors(collect_embedding_texts(df, selected_rubric))

    questions = selected_rubric.get("questions", [])
    score_rows = []
    conf_rows = []
    feedback_rows = []
    semantic_pattern_rows = []
    traceability_rows = []

    accepted_count = 0
    caution_count = 0
    review_count = 0
    total_answers = 0

    question_stats = {}
    for p in questions:
        pid = p.get("id")
        question_stats[pid] = {
            "scores": [],
            "confidences": [],
            "accepted": 0,
            "review": 0,
            "caution": 0,
            "total": 0
        }

    for _, row in df.iterrows():
        sid = get_student_id(row, df)
        nombre = get_student_name(row, df)
        score_row = {"student_id": sid, "nombre": nombre}
        conf_row = {"student_id": sid, "nombre": nombre}
        total = 0.0

        for p in questions:
            pid = p["id"]
            col = find_item_column(df, pid)
            answer = row.get(col, "") if col else ""

            score, conf, fb, status = score_answer(answer, p)

            score_row[f"{pid}_score"] = score
            conf_row[f"{pid}_confidence"] = conf
            total += score
            total_answers += 1

            question_stats[pid]["scores"].append(score)
            question_stats[pid]["confidences"].append(conf)
            question_stats[pid]["total"] += 1

            if status == "aceptado":
                accepted_count += 1
                question_stats[pid]["accepted"] += 1
            elif status == "aceptado_con_cautela":
                caution_count += 1
                question_stats[pid]["caution"] += 1
            else:
                review_count += 1
                question_stats[pid]["review"] += 1

            diagnosis = semantic_diagnosis(answer, p, score=score, confidence=conf, status=status)
            semantic_pattern_rows.append({
                "student_id": sid,
                "nombre": nombre,
                "pregunta_id": pid,
                "tipo_item": display_item_type(p.get("item_type", "")),
                "score": score,
                "max_score": p.get("max_score", ""),
                "confidence": conf,
                "status": status,
                **diagnosis
            })
            traceability_rows.append(build_traceability_row(sid, nombre, pid, answer, score, conf, status, diagnosis, fb))

            feedback_rows.append({
                "student_id": sid,
                "nombre": nombre,
                "rubric": selected_rubric.get("_rubric_display_name", selected_rubric.get("name", "")),
                "pregunta_id": pid,
                "columna_detectada": col,
                "prompt": p.get("prompt", ""),
                "answer": answer,
                "score": score,
                "max_score": p.get("max_score", ""),
                "confidence": conf,
                "status": status,
                "feedback": fb
            })

        total_score = float(selected_rubric.get("total_score", 0)) or sum(float(p.get("max_score", 0)) for p in questions) or 1
        score_row["total"] = round(total, 2)
        score_row["porcentaje"] = round((total / total_score) * 100, 2)
        score_row["nivel_desempeno"] = performance_level(score_row["porcentaje"])
        score_rows.append(score_row)
        conf_rows.append(conf_row)

    rubric_name = selected_rubric.get("_rubric_display_name", selected_rubric.get("name", "Rúbrica"))
    insights_rows, problematic_questions = build_question_insights(question_stats, questions)
    type_insights_rows = build_type_insights(insights_rows)
    interpretation = build_interpretation(insights_rows, problematic_questions, len(df), rubric_name)

    output_name = f"evalia_resultados_{Path(file.filename).stem}_{Path(selected_rubric.get('_rubric_filename', 'rubrica')).stem}.xlsx"
    output_path = OUTPUT_DIR / output_name

    auto_rate = round((accepted_count / total_answers) * 100, 1) if total_answers else 0
    review_rate = round((review_count / total_answers) * 100, 1) if total_answers else 0
    caution_rate = round((caution_count / total_answers) * 100, 1) if total_answers else 0

    summary_rows = [{
        "rubric": rubric_name,
        "rubric_file": selected_rubric.get("_rubric_filename", ""),
        "students_processed": len(df),
        "questions_evaluated": len(questions),
        "total_answers": total_answers,
        "accepted": accepted_count,
        "accepted_with_caution": caution_count,
        "review_required": review_count,
        "alta_confianza_pct": auto_rate,
        "parcial_intermedia_pct": caution_rate,
        "revision_sugerida_pct": review_rate
    }]

    teacher_report_rows = build_teacher_report_rows(
        score_rows=score_rows,
        insights_rows=insights_rows,
        interpretation=interpretation,
        problematic_questions=problematic_questions,
        rubric_name=rubric_name,
        total_students=len(df),
        total_questions=len(questions),
        auto_rate=auto_rate,
        caution_rate=caution_rate,
        review_rate=review_rate
    )

    elapsed_seconds = round(time.time() - started_at, 3)
    log_event("upload_completed", output=output_name, seconds=elapsed_seconds, accepted_pct=auto_rate, caution_pct=caution_rate, review_pct=review_rate)

    embeddings_rows = [{
        "version": APP_VERSION,
        "embeddings_solicitados": EMBEDDING_STATUS.get("enabled_requested"),
        "embeddings_disponibles": EMBEDDING_STATUS.get("available"),
        "modo": EMBEDDING_STATUS.get("mode"),
        "modelo": EMBEDDING_STATUS.get("model"),
        "dispositivo": EMBEDDING_STATUS.get("device"),
        "mensaje": EMBEDDING_STATUS.get("message"),
        "batch_status": embedding_batch_info.get("status"),
        "batch_precomputed_items": embedding_batch_info.get("precomputed"),
        "batch_seconds": embedding_batch_info.get("seconds"),
        "nota": "Si embeddings_disponibles=False, Evalia procesó normalmente con reglas semánticas. Para activar embeddings reales agregue sentence-transformers, torch y EVALIA_EMBEDDINGS=1."
    }]

    technical_trace_rows = [{
        "version": APP_VERSION,
        "archivo_respuestas": file.filename,
        "rubrica": rubric_name,
        "estudiantes": len(df),
        "preguntas": len(questions),
        "respuestas_evaluadas": total_answers,
        "tiempo_procesamiento_segundos": elapsed_seconds,
        "cache_semantico_items": len(SEMANTIC_CACHE),
        "embedding_batch_status": embedding_batch_info.get("status"),
        "embedding_batch_precomputed_items": embedding_batch_info.get("precomputed"),
        "embedding_batch_seconds": embedding_batch_info.get("seconds"),
        "embedding_mode": EMBEDDING_STATUS.get("mode"),
        "embedding_model": EMBEDDING_STATUS.get("model"),
        "embedding_device": EMBEDDING_STATUS.get("device"),
        "log_runtime": str(LOG_PATH.name),
        "criterio_robustez": "baseline congelado + validación preventiva + logging + caché semántico + precálculo batch de embeddings + trazabilidad por respuesta"
    }]

    validation_science_rows = [{
        "version": APP_VERSION,
        "modo_validacion": "baseline_prevalidacion",
        "baseline_congelado": True,
        "archivo_respuestas": file.filename,
        "rubrica": rubric_name,
        "estudiantes": len(df),
        "preguntas": len(questions),
        "respuestas_evaluadas": total_answers,
        "alta_confianza_pct": auto_rate,
        "parcial_intermedia_pct": caution_rate,
        "revision_sugerida_pct": review_rate,
        "tiempo_total_segundos": elapsed_seconds,
        "embedding_mode": EMBEDDING_STATUS.get("mode"),
        "embedding_model": EMBEDDING_STATUS.get("model"),
        "embedding_batch_status": embedding_batch_info.get("status"),
        "embedding_batch_precomputed_items": embedding_batch_info.get("precomputed"),
        "embedding_batch_seconds": embedding_batch_info.get("seconds"),
        "cache_semantico_items": len(SEMANTIC_CACHE),
        "variables_exportadas": "score, confidence, status, feedback, patrones semánticos, trazabilidad, embeddings, robustez técnica",
        "uso_recomendado": "Usar esta versión como baseline estable para comparar Evalia contra corrección humana en corpus reales."
    }]

    # Se agrega una fila de identidad de producto al reporte docente.
    teacher_report_rows.insert(0, {
        "seccion": "Producto",
        "indicador": "Sistema",
        "valor": "Evalia by Altiora · Inteligencia Semántica Docente · v3.5 baseline validación"
    })

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(teacher_report_rows).to_excel(writer, sheet_name="REPORTE_DOCENTE", index=False)
        pd.DataFrame(score_rows).to_excel(writer, sheet_name="Puntajes", index=False)
        pd.DataFrame(conf_rows).to_excel(writer, sheet_name="Confianza", index=False)
        pd.DataFrame(feedback_rows).to_excel(writer, sheet_name="Feedback", index=False)
        pd.DataFrame(semantic_pattern_rows).to_excel(writer, sheet_name="PATRONES_SEMANTICOS", index=False)
        pd.DataFrame(traceability_rows).to_excel(writer, sheet_name="TRAZABILIDAD_EVALIA", index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Resumen_Tecnico", index=False)
        pd.DataFrame(technical_trace_rows).to_excel(writer, sheet_name="ROBUSTEZ_TECNICA", index=False)
        pd.DataFrame(embeddings_rows).to_excel(writer, sheet_name="EMBEDDINGS", index=False)
        pd.DataFrame(validation_science_rows).to_excel(writer, sheet_name="VALIDACION_CIENTIFICA", index=False)
        pd.DataFrame(insights_rows).to_excel(writer, sheet_name="Analisis_Items", index=False)
        pd.DataFrame(type_insights_rows).to_excel(writer, sheet_name="Analisis_Tipos", index=False)
        pd.DataFrame([{
            "rubric": rubric_name,
            "resumen_interpretativo": interpretation,
            "preguntas_potencialmente_problematicas": ", ".join(problematic_questions) if problematic_questions else "Sin preguntas críticas"
        }]).to_excel(writer, sheet_name="Interpretacion", index=False)

        format_workbook(writer)

    problematic_display = ", ".join(problematic_questions) if problematic_questions else "Sin preguntas críticas"

    return HTMLResponse(
        f"""
        <!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Resultados · Evalia</title>{base_css()}</head>
        <body><div class="page"><main class="shell">
          {shell_topbar("Reporte generado · Evalia by Altiora", "CRB Engine · v3.5 baseline validación")}
          <section class="result-card">
            <h1>Procesamiento completado</h1>
            <p class="lead">Evalia aplicó la rúbrica <strong>{escape(rubric_name)}</strong> y generó un reporte Excel explicable.</p>
            <div class="metric-grid">
              <div class="metric"><div class="metric-value">{len(df)}</div><div class="metric-label">estudiante(s)</div></div>
              <div class="metric"><div class="metric-value">{auto_rate}%</div><div class="metric-label">alta confianza</div></div>
              <div class="metric"><div class="metric-value">{caution_rate}%</div><div class="metric-label">parcial/intermedia</div></div>
              <div class="metric"><div class="metric-value">{review_rate}%</div><div class="metric-label">revisión sugerida</div></div>
            </div>
            <p class="hint">Los tres estados de respuesta suman 100%: alta confianza, parcial/intermedia y revisión sugerida.</p>
            <p class="lead"><strong>Insight inicial:</strong> {escape(problematic_display)}</p>
            <div class="actions">
              <a class="button" href="/download/{output_name}">Descargar reporte Excel</a>
              <a class="button secondary" href="/">Evaluar otro archivo</a>
            </div>
          </section>
          {footer_altiora()}
        </main></div></body></html>
        """
    )


@app.get("/download/{filename}")
def download(filename: str):
    path = OUTPUT_DIR / filename
    if not path.exists():
        return HTMLResponse(
            f"""
            <!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Archivo no encontrado · Evalia</title>{base_css()}</head>
            <body><div class="page"><main class="shell">{shell_topbar("Archivo no encontrado", "CRB Engine · v3.5 baseline validación")}<div class="result-card">
              <h1>No se pudo acceder al archivo</h1>
              <div class="error">El reporte solicitado no existe o no fue generado correctamente.</div>
              <p class="lead">Vuelve al inicio y procesa nuevamente la rúbrica y las respuestas.</p>
              <br><a class="button" href="/">Volver</a>
            </div>{footer_altiora()}</main></div></body></html>
            """
        )

    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if filename.lower().endswith(".html"):
        media_type = "text/html"

    return FileResponse(
        path,
        media_type=media_type,
        filename=filename
    )
